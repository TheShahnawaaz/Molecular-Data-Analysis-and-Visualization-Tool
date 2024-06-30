# import tkinter
# import tkinter.messagebox
import customtkinter
from customtkinter import filedialog
from customtkinter import END
from typing import Callable, Union
import os
from PIL.Image import open as open_PIL
from PIL.ImageTk import PhotoImage
from pickle import load, dump
from CTkMessagebox import CTkMessagebox, ProgressCTkMessagebox
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.drawing.image import Image
import sys
import random

# theme = "System"


if getattr(sys, 'frozen', False):
    # we are running in a bundle (exe)
    base_path = sys._MEIPASS
    import pyi_splash
else:
    # we are running in a normal Python environment
    base_path = os.path.abspath('.')


def add_cell_value(worksheet, row_num, col_num, value, **cell_params):
    """
    Adds a value to a cell in a worksheet and sets its properties.

    Args:
        worksheet: An openpyxl.worksheet.worksheet.Worksheet object.
        row_num (int): The row number of the cell.
        col_num (int): The column number of the cell.
        value: The value to add to the cell.
        cell_params: Any additional keyword arguments will be used as properties of the cell (e.g. fill, font, border, etc.).
    """

    cell = worksheet.cell(row=row_num, column=col_num, value=value)
    for param, value in cell_params.items():
        setattr(cell, param, value)


# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("System")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("blue")


class CustomSpinbox(customtkinter.CTkFrame):
    def __init__(self, *args,
                 width: int = 100,
                 height: int = 32,
                 step_size: Union[int, float] = 1,
                 data_type: type = float,
                 command: Callable = None,
                 **kwargs):
        super().__init__(*args, width=width, height=height, **kwargs)

        self.step_size = step_size
        self.data_type = data_type
        self.command = command

        self.configure(fg_color=("gray78", "gray28"))  # set frame color

        self.grid_columnconfigure((0, 2), weight=0)  # buttons don't expand
        self.grid_columnconfigure(1, weight=1)  # entry expands

        self.subtract_button = customtkinter.CTkButton(self, text="-", width=height-6, height=height-6,
                                                       command=self.subtract_button_callback)
        self.subtract_button.grid(row=0, column=0, padx=(3, 0), pady=3)

        self.entry = customtkinter.CTkEntry(
            self, width=width-(2*height), height=height-6, border_width=0)
        self.entry.grid(row=0, column=1, columnspan=1,
                        padx=3, pady=3, sticky="ew")

        self.add_button = customtkinter.CTkButton(self, text="+", width=height-6, height=height-6,
                                                  command=self.add_button_callback)
        self.add_button.grid(row=0, column=2, padx=(0, 3), pady=3)

        # default value
        self.entry.insert(0, "0")

    def add_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = self.data_type(self.entry.get()) + self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            msg = CTkMessagebox(
                title="Error", message="Invalid value " + str(self.entry.get()), icon="cancel", master=app)
            return

    def subtract_button_callback(self):
        if self.command is not None:
            self.command()
        try:
            value = self.data_type(self.entry.get()) - self.step_size
            self.entry.delete(0, "end")
            self.entry.insert(0, value)
        except ValueError:
            msg = CTkMessagebox(
                title="Error", message="Invalid value "+str(self.entry.get()), icon="cancel", master=app)
            return

    def get(self) -> Union[float, int, None]:
        try:
            return self.data_type(self.entry.get())
        except ValueError:
            msg = CTkMessagebox(
                title="Error", message="Invalid value "+str(self.entry.get()), icon="cancel", master=app)
            print("here")
            return None

    def set(self, value: Union[float, int]):
        self.entry.delete(0, "end")
        self.entry.insert(0, str(self.data_type(value)))

class HelpWindow(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("900x700")
        self.title("Help Window")
        self.resizable(False, False)
        self.iconbitmap(os.path.join(base_path,"assets", "icon.ico"))
        self.attributes("-topmost", True)
        self.segemented_button_var = customtkinter.StringVar(value="Home")

        self.top_button_label = ["Home", "Generate FALDI-pops.sum", "Molecular Inputs", "References"]

        self.frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.frame.pack(fill="both", expand=True)

        # self.progressbar = customtkinter.CTkProgressBar(self, orientation="horizontal", mode="determinate",determinate_speed=0.01)
        # self.progressbar.pack(fill="x", padx=30, pady=(10, 0))
        self.next_button = customtkinter.CTkButton(self, text="\u2192", command=self.next, width=20)
        self.next_button.pack(side="right", padx=30, pady=10)
        self.previous_button = customtkinter.CTkButton(self, text="\u2190", command=self.previous, width=20)
        self.previous_button.pack(side="left", padx=30, pady=10)
        self.segmendted_button = customtkinter.CTkSegmentedButton(self, values=self.top_button_label, command=self.handle_segmented_button, variable=self.segemented_button_var,font=customtkinter.CTkFont(family="Candara",size=16))
        self.segmendted_button.pack(fill="x", padx=30, pady=10)


        self.frames = []
        self.frame_idx = 0
        self.frames.append(self.Frame1(self.frame))
        self.frames.append(self.Frame2(self.frame))
        self.frames.append(self.Frame3(self.frame))
        self.frames.append(self.Frame4(self.frame))
        # self.progressbar.set(1/len(self.frames)-1/10000)

        self.show_frame(self.frame_idx)

        self.bind("<Right>", lambda event: self.next())
        self.bind("<Left>", lambda event: self.previous())
        self.focus()

    def handle_segmented_button(self, value):
        idx = self.top_button_label.index(value)
        self.frame_idx = idx
        self.show_frame(idx)
    
    def show_frame(self, frame_idx: int):
        self.segemented_button_var.set(self.top_button_label[frame_idx])
        if hasattr(self, "current_frame"):
            self.current_frame.pack_forget()
        self.current_frame = self.frames[frame_idx]
        self.frames[frame_idx].pack(fill="both", expand=True)

    def next(self):
        
        self.frame_idx = (self.frame_idx+1) % len(self.frames)
        
        self.show_frame(self.frame_idx)
        
        # self.progressbar.set((self.progressbar.get()+(1/len(self.frames)))%1)

    def previous(self):
        self.frame_idx = (self.frame_idx-1) % len(self.frames)
        self.show_frame(self.frame_idx)

        # self.progressbar.set((self.progressbar.get()-(1/len(self.frames)))%1)

    class Frame1(customtkinter.CTkFrame):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            i = open_PIL(os.path.join(base_path,"assets", "Screen-1.jpg"))
            self.image1 = customtkinter.CTkImage(light_image= i, dark_image=i, size=(800,600))
            self.label = customtkinter.CTkLabel(self, text="",image=self.image1)
            self.label.pack(fill="both", expand=True)



    class Frame2(customtkinter.CTkFrame):

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            i = open_PIL(os.path.join(base_path,"assets", "Screen-2.jpg"))
            self.image2 = customtkinter.CTkImage(light_image= i, dark_image=i, size=(800,600))
            self.label = customtkinter.CTkLabel(self, text="",image=self.image2)
            self.label.pack(fill="both", expand=True)
                                                 
    class Frame3(customtkinter.CTkFrame):
            
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            i = open_PIL(os.path.join(base_path,"assets", "Screen-3.jpg"))
            self.image3 = customtkinter.CTkImage(light_image= i, dark_image=i, size=(800,600))
            self.label = customtkinter.CTkLabel(self, text="",image=self.image3)
            self.label.pack(fill="both", expand=True)

    class Frame4(customtkinter.CTkFrame):
            
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)

            self.columnconfigure(0, weight=1)
            self.rowconfigure(0, weight=1)
            self.text_box_for_references = customtkinter.CTkTextbox(self, height=550, font=customtkinter.CTkFont(family="Candara",size=18))

            self.text_box_for_references.grid(row=0, column=0, padx=80, pady=(0, 0), sticky="ew")
            self.text_box_for_references.insert("0.0", """Suggested literature

I. Cukrowski, ‘A unified molecular-wide and electron density based concept of chemical bonding’, WIREs Comput. Mol. Sci. 2021; e1579. doi.org/10.1002/wcms.1579.

T. G. Bates, J. H. de Lange and I. Cukrowski, ‘The CH⋅⋅⋅HC interaction in biphenyl is a delocalized, molecular-wide and entirely non-classical interaction: results from FALDI analysis.’ J. Comput. Chem., 2021, 42, 706-718. DOI: 10.1002/jcc.26491

J. H. de Lange, D. M. E. van Niekerk and I. Cukrowski, “FALDI-based decomposition of an atomic interaction line leads to 3D representation of the multicentre nature of interactions”, J. Comput. Chem., 39 (2018) 973–985. DOI: 10.1002/jcc.25175

J. H. de Lange, D. M. E. van Niekerk and I. Cukrowski, “FALDI-Based Criterion for and the Origin of an Electron Density Bridge with an Associated (3,–1) Critical Point on Bader’s Molecular Graph”, J. Comput. Chem., 39 (2018) 2283-2299. DOI:10.1002/jcc.25548

J. H. de Lange and I. Cukrowski, “Exact and Exclusive Electron Localization Indices Within QTAIM Atomic Basins”, J. Comput. Chem., 39 (2018) 1517–1530. DOI: 10.1002/jcc.25223

J. H. de Lange and I. Cukrowski, “Toward Deformation Densities for Intramolecular Interactions without Radical Reference States Using the Fragment, Atom, Localized, Delocalized, and Interatomic (FALDI) Charge Density Decomposition Scheme”, J. Comput. Chem. 38 (2017) 981–997. DOI: 10.1002/jcc.24772""")
            self.text_box_for_references.configure(state="disabled")


class FaldiWindow(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("320x300+300+300")
        self.master = args[0]
        self.title("FALDI Advanced Features")
        self.resizable(True, True)
        self.iconbitmap(os.path.join(base_path, "assets", "icon.ico"))
        self.attributes("-topmost", True)
        self.segemented_button_var = customtkinter.StringVar(value="Home")
        # Make the 0th column expandable
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        


        # add a Label and a button in the same line 
        self.top_label = customtkinter.CTkLabel(self, text="Enter custom commands for Faldi:", font=customtkinter.CTkFont(family="Candara"))
        self.top_label.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="w")
        self.top_button = customtkinter.CTkButton(self, text="\u2753", command=self.open_help, border_width=0, width=30)
        self.top_button.grid(row=0, column=1, padx=20, pady=(20, 10), sticky="ew")

        # add a text box
        self.text_box = customtkinter.CTkTextbox(self, font=customtkinter.CTkFont(family="Candara",size=15))
        self.text_box.grid(row=1, column=0, padx=20, pady=(0, 10), sticky="nsew", columnspan=2)
        
        self.submit_button = customtkinter.CTkButton(self, text="Submit", command=self.submit, border_width=0)
        self.submit_button.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="nsew", columnspan=2)


    def open_help(self):
        help_doc_path = os.path.join(base_path, "assets", "FALDIAdvancedOptionsDescription.pdf")
        os.startfile(help_doc_path)
    
    def submit(self):
        self.master.faldi_commands = self.text_box.get("0.0", END).strip()
        self.destroy()

class ATOMIC_App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Molecular-wide electron (de)localization atomic counts")
        self.geometry(f"{800}x{600}")
        self.resizable(True, True)
        self.minsize(800, 600)
        self.iconbitmap(os.path.join(base_path,"assets", "icon.ico"))



        # set minimum size
        self.data_for_this_file = None
        self.data_for_this_file_image = None

        help_key = "<Control-h>"
        self.help_window = None
        self.bind(help_key, func=lambda event: self.current_frame.open_help())

        faldi_key = "<Control-f>"
        self.faldi_window = None
        self.faldi_commands = None
        self.bind(faldi_key, func=lambda event: self.open_faldi_command())

        # # configure grid layout (4x4)
        # self.grid_columnconfigure(1, weight=1)
        # self.grid_columnconfigure((2, 3), weight=0)
        # self.grid_rowconfigure((0, 1, 2), weight=1)

        # # create sidebar frame with widgets
        # self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        # self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        # self.sidebar_frame.grid_rowconfigure(4, weight=1)
        # self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="CustomTkinter", font=customtkinter.CTkFont(size=20, weight="bold"))
        # self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        # self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        # self.sidebar_button_1.grid(row=1, column=0, padx=20, pady=10)
        # self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        # self.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
        # self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, command=self.sidebar_button_event)
        # self.sidebar_button_3.grid(row=3, column=0, padx=20, pady=10)
        # self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        # self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        # self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],command=self.change_appearance_mode_event)
        # self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))

        # self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        # self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        # self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
        #                                                        command=self.change_scaling_event)
        # self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        self.container = customtkinter.CTkFrame(self, corner_radius=0)
        self.container.pack(fill="both", expand=True)

        self.home_frame = ATOMIC_HomeFrame(self.container, self)
        self.input_frame = ATOMIC_InputFrame(self.container, self)
        self.page2_frame = ATOMIC_Page2Frame(self.container, self)
        self.page3_frame = ATOMIC_Page3Frame(self.container, self)

        # self.home_frame.pack(fill="both", expand=True)
        # self.setting_frame.pack(fill="both", expand=True)

        # Show the home frame by default
        # self.home_frame.tkraise()
        self.show_frame(self.home_frame)
        

    def open_help(self, idx):
        if self.help_window is None or not self.help_window.winfo_exists():
            self.help_window = HelpWindow(self)
        
        for i in range(idx):
            self.help_window.next()
        self.help_window.focus()
    
    def open_faldi_command(self):
        if self.current_frame != self.page2_frame:
            return
        if self.faldi_window is None or not self.faldi_window.winfo_exists():
            self.faldi_window = FaldiWindow(self)



    def show_frame(self, frame):

        if hasattr(self, "current_frame"):
            self.current_frame.pack_forget()

        frame.pack(fill="both", expand=True)
        self.current_frame = frame

        # self.color_theme_label = customtkinter.CTkLabel(self.sidebar_frame, text="Color Theme:", anchor="w")
        # self.color_theme_label.grid(row=9, column=0, padx=20, pady=(10, 0))
        # self.color_theme_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["blue", "green", "dark-blue"], command=self.change_color_theme_event)
        # self.color_theme_optionemenu.grid(row=10, column=0, padx=20, pady=(10, 10))

        # # create main entry and button
        # self.entry = customtkinter.CTkEntry(self, placeholder_text="CTkEntry")
        # self.entry.grid(row=3, column=1, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")

        # self.main_button_1 = customtkinter.CTkButton(self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"))
        # self.main_button_1.grid(row=3, column=3, padx=(20, 20), pady=(20, 20), sticky="nsew")

        # # create textbox
        # self.textbox = customtkinter.CTkTextbox(self, width=250)
        # self.textbox.grid(row=0, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")

        # # create tabview
        # self.tabview = customtkinter.CTkTabview(self, width=250)
        # self.tabview.grid(row=0, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        # self.tabview.add("CTkTabview")
        # self.tabview.add("Tab 2")
        # self.tabview.add("Tab 3")
        # self.tabview.tab("CTkTabview").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        # self.tabview.tab("Tab 2").grid_columnconfigure(0, weight=1)

        # self.optionmenu_1 = customtkinter.CTkOptionMenu(self.tabview.tab("CTkTabview"), dynamic_resizing=False,values=["Value 1", "Value 2", "Value Long Long Long"])
        # self.optionmenu_1.grid(row=0, column=0, padx=20, pady=(20, 10))
        # self.combobox_1 = customtkinter.CTkComboBox(self.tabview.tab("CTkTabview"), values=["Value 1", "Value 2", "Value Long....."])
        # self.combobox_1.grid(row=1, column=0, padx=20, pady=(10, 10))
        # self.string_input_button = customtkinter.CTkButton(self.tabview.tab("CTkTabview"), text="Open CTkInputDialog",
        #                                                    command=self.open_input_dialog_event)
        # self.string_input_button.grid(row=2, column=0, padx=20, pady=(10, 10))
        # self.label_tab_2 = customtkinter.CTkLabel(self.tabview.tab("Tab 2"), text="CTkLabel on Tab 2")
        # self.label_tab_2.grid(row=0, column=0, padx=20, pady=20)

        # # create radiobutton frame
        # self.radiobutton_frame = customtkinter.CTkFrame(self)
        # self.radiobutton_frame.grid(row=0, column=3, padx=(20, 20), pady=(20, 0), sticky="nsew")
        # self.radio_var = tkinter.IntVar(value=0)
        # self.label_radio_group = customtkinter.CTkLabel(master=self.radiobutton_frame, text="CTkRadioButton Group:")
        # self.label_radio_group.grid(row=0, column=2, columnspan=1, padx=10, pady=10, sticky="")
        # self.radio_button_1 = customtkinter.CTkRadioButton(master=self.radiobutton_frame,variable=self.radio_var, value=0, text="New Label")
        # self.radio_button_1.grid(row=1, column=2, pady=10, padx=20, sticky="n")

        # self.radio_button_2 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, variable=self.radio_var, value=1)
        # self.radio_button_2.grid(row=2, column=2, pady=10, padx=20, sticky="n")

        # self.radio_button_3 = customtkinter.CTkRadioButton(master=self.radiobutton_frame, variable=self.radio_var, value=2)
        # self.radio_button_3.grid(row=3, column=2, pady=10, padx=20, sticky="n")

        # # create slider and progressbar frame
        # self.slider_progressbar_frame = customtkinter.CTkFrame(self, fg_color="transparent")
        # self.slider_progressbar_frame.grid(row=1, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")
        # self.slider_progressbar_frame.grid_columnconfigure(0, weight=1)
        # self.slider_progressbar_frame.grid_rowconfigure(4, weight=1)
        # self.seg_button_1 = customtkinter.CTkSegmentedButton(self.slider_progressbar_frame)
        # self.seg_button_1.grid(row=0, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
        # self.progressbar_1 = customtkinter.CTkProgressBar(self.slider_progressbar_frame)
        # self.progressbar_1.grid(row=1, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
        # self.progressbar_2 = customtkinter.CTkProgressBar(self.slider_progressbar_frame)
        # self.progressbar_2.grid(row=2, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
        # self.slider_1 = customtkinter.CTkSlider(self.slider_progressbar_frame, from_=0, to=1, number_of_steps=4)
        # self.slider_1.grid(row=3, column=0, padx=(20, 10), pady=(10, 10), sticky="ew")
        # self.slider_2 = customtkinter.CTkSlider(self.slider_progressbar_frame, orientation="vertical")
        # self.slider_2.grid(row=0, column=1, rowspan=5, padx=(10, 10), pady=(10, 10), sticky="ns")
        # self.progressbar_3 = customtkinter.CTkProgressBar(self.slider_progressbar_frame, orientation="vertical")
        # self.progressbar_3.grid(row=0, column=2, rowspan=5, padx=(10, 20), pady=(10, 10), sticky="ns")

        # # create scrollable frame
        # self.scrollable_frame = customtkinter.CTkScrollableFrame(self, label_text="CTkScrollableFrame")
        # self.scrollable_frame.grid(row=1, column=2, padx=(20, 0), pady=(20, 0), sticky="nsew")
        # self.scrollable_frame.grid_columnconfigure(0, weight=1)
        # self.scrollable_frame_switches = []
        # for i in range(100):
        #     switch = customtkinter.CTkSwitch(master=self.scrollable_frame, text=f"CTkSwitch {i}")
        #     switch.grid(row=i, column=0, padx=10, pady=(0, 20))
        #     self.scrollable_frame_switches.append(switch)

        # # create checkbox and switch frame
        # self.checkbox_slider_frame = customtkinter.CTkFrame(self)
        # self.checkbox_slider_frame.grid(row=1, column=3, padx=(20, 20), pady=(20, 0), sticky="nsew")
        # self.checkbox_1 = customtkinter.CTkCheckBox(master=self.checkbox_slider_frame)
        # self.checkbox_1.grid(row=1, column=0, pady=(20, 0), padx=20, sticky="n")
        # self.checkbox_2 = customtkinter.CTkCheckBox(master=self.checkbox_slider_frame)
        # self.checkbox_2.grid(row=2, column=0, pady=(20, 0), padx=20, sticky="n")
        # self.checkbox_3 = customtkinter.CTkCheckBox(master=self.checkbox_slider_frame)
        # self.checkbox_3.grid(row=3, column=0, pady=20, padx=20, sticky="n")

        # # set default values
        # self.sidebar_button_3.configure(state="disabled", text="Disabled CTkButton")
        # self.checkbox_3.configure(state="disabled")
        # self.checkbox_1.select()
        # self.scrollable_frame_switches[0].select()
        # self.scrollable_frame_switches[4].select()
        # self.radio_button_3.configure(state="disabled")
        # self.appearance_mode_optionemenu.set("System")
        # self.scaling_optionemenu.set("100%")

        # self.optionmenu_1.set("CTkOptionmenu")
        # self.combobox_1.set("CTkComboBox")
        # self.slider_1.configure(command=self.progressbar_2.set)
        # self.slider_2.configure(command=self.progressbar_3.set)
        # self.progressbar_1.configure(mode="indeterminnate")
        # self.progressbar_1.start()
        # self.textbox.insert("0.0", "CTkTextbox\n\n" + "Lorem ipsum dolor sit amet, consetetur sadipscing elitr, sed diam nonumy eirmod tempor invidunt ut labore et dolore magna aliquyam erat, sed diam voluptua.\n\n" * 20)
        # self.seg_button_1.configure(values=["CTkSegmentedButton", "Value 2", "Value 3"])
        # self.seg_button_1.set("Value 2")

    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

        self.home_frame.appearance_mode_optionemenu.set(new_appearance_mode)
        self.input_frame.appearance_mode_optionemenu.set(new_appearance_mode)
        self.page2_frame.appearance_mode_optionemenu.set(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)
        print(self.title)

        self.home_frame.scaling_optionemenu.set(new_scaling)
        self.input_frame.scaling_optionemenu.set(new_scaling)
        self.page2_frame.scaling_optionemenu.set(new_scaling)


class ATOMIC_HomeFrame(customtkinter.CTkFrame):
    def __init__(self, master, container, **kwargs):
        super().__init__(master, **kwargs)



        self.container = container

        # configure grid layout (4x4)
        self.grid_columnconfigure((2, 3), weight=1)
        # self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((3), weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="MOWeD-LAC", font=customtkinter.CTkFont(family="Candara",size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.appearance_mode_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=[
                                                                       "Light", "Dark", "System"], command=self.container.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="GUI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%", "130%", "140%", "150%"],
                                                               command=self.container.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # Default values

        self.appearance_mode_optionemenu.set(
            customtkinter.get_appearance_mode())
        self.scaling_optionemenu.set("100%")


        self.logo_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.logo_frame.grid(row=0, column=1, padx=20, pady=(150, 0), columnspan=3)


        self.logo_image = customtkinter.CTkImage(light_image=open_PIL(os.path.join(base_path,"assets", "icon.ico")), dark_image=open_PIL(os.path.join(base_path,"assets", "icon.ico")), size=(40, 40))
        self.name_label = customtkinter.CTkLabel(
            self.logo_frame, text="MOWeD-LAC", font=customtkinter.CTkFont(family="Candara",size=50, weight="bold"), anchor="w")
        self.name_label.grid(row=0, column=1, padx=0,
                             pady=(0, 0), columnspan=2)
        self.logo_label = customtkinter.CTkLabel(
            self.logo_frame,text="", image=self.logo_image, fg_color="transparent")
        self.logo_label.grid(row=0, column=0, padx=0,
                             pady=(0, 0), columnspan=1)

        self.des_label = customtkinter.CTkLabel(
            self, text="Molecular-wide electron (de)localization atomic counts", font=customtkinter.CTkFont(family="Verdana",size=12))
        self.des_label.grid(row=1, column=1, pady=(0, 10), columnspan=3)

        self.contributors_label = customtkinter.CTkTextbox(
            self, border_width=0, fg_color="transparent", font=customtkinter.CTkFont(family="Courier New",size=12),width=480)
        self.contributors_label.grid(
            row=2, column=1, padx=(0, 0), pady=(50, 0), columnspan=3)
        contributors = "Prof Ignacy Cukrowski: Concept, functionality and design"
        contributors += "\nemail: ignacy.cukrowski@up.ac.za"
        contributors += "\n\nMr Shahnawaz Hussain: Implementation and coding of MOWeD-LAC"
        contributors += "\nemail: mowed.shahnawaz@gmail.com"
        contributors += "\n\nDr Jurgens H. de Lange: Implementation and coding of FALDI"
        contributors += "\nemail: jurgens.delange@up.ac.za"
        self.contributors_label.insert("0.0", contributors)
        self.contributors_label.configure(state="disabled")



        # self
          
        self.faldi_3d_button = customtkinter.CTkButton(
            self, text="FALDI 3D", command=self.faldi_3d, border_width=0)
        self.faldi_3d_button.grid(
            row=4, column=2,  columnspan=2, padx=(0,20) ,
            pady=(0, 10), sticky="nsew")
        

        self.help_button = customtkinter.CTkButton(
            self, text="\u2753", command=self.open_help, border_width=0, width=30)
        self.help_button.grid(row=5, column=1, padx=20,
                                pady=(0, 20))
        


        self.generate_sum_button = customtkinter.CTkButton(
            self, text="Generate FALDI-pops.sum file", command=self.generate_sum_file, border_width=0)
        self.generate_sum_button.grid(
            row=5, column=2, padx=0, pady=(0, 20), sticky="nsew")
      

        self.submit_button = customtkinter.CTkButton(
            self, text="Choose FALDI-pops.sum file", command=self.convert_sum_file, border_width=0)
        self.submit_button.grid(row=5, column=3, padx=20,
                                pady=(0, 20), sticky="nsew")

        # # create main entry and button
        # self.file_name_entry = customtkinter.CTkEntry(
        #     self, placeholder_text="Select a file", corner_radius=50, border_width=1)
        # self.file_name_entry.grid(row=4, column=1, columnspan=2, padx=(
        #     20, 0), pady=(20, 20), sticky="nsew")

        # self.browse_button = customtkinter.CTkButton(
        #     self, text="Browse",  border_width=1, command=lambda: self.handle_browse_button_click(self.file_name_entry))
        # self.browse_button.grid(row=4, column=3, padx=(
        #     10, 10), pady=(20, 20), sticky="nsew")

    def faldi_3d(self):
        print("FALDI 3D")
        self.container.show_frame(self.container.page3_frame)

    def open_help(self):
        self.container.open_help(0)

    def generate_sum_file(self):
        print("Generate SUM file")
        self.container.show_frame(self.container.page2_frame)

    def convert_sum_file(self):
        print("Convert SUM file")
        sum_filename = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            initialfile="FALDI-pops.sum",
            title="Select a SUM file",
            filetypes=[("SUM files", "*.sum")]
        )

        if not sum_filename:
            msg = CTkMessagebox(
                title="Warning!!", message="No SUM file selected", icon="warning", master=app)
            if msg.get() == "OK":
                pass
            return
        try:
            dir_name, _ = os.path.split(sum_filename)

            os.chdir(dir_name)
            with open(sum_filename, 'r') as sum_file:
                sum_data = sum_file.readlines()
            # Split each line of the SUM file by spaces to separate the columns
            split_data = [line.split() for line in sum_data]

            print("FALDI-pops.sum file read successfully")

            # create a new workbook object
            workbook = openpyxl.Workbook()

            # select the active worksheet
            worksheet = workbook.active

            # write the data to the worksheet
            for row in split_data:
                worksheet.append(row)

            # # Ask the user to select where to save the output file
            save_filename = "example.xlsx"

            # # save the workbook to a file
            # workbook.save(save_filename)

            # # Load workbook and select the worksheet
            # workbook = openpyxl.load_workbook('example.xlsx')
            # worksheet = workbook.active

            # Find the index of the row that separates the two tables
            for i, row in enumerate(worksheet.iter_rows()):
                if all(cell.value is None for cell in row):
                    separator_index = i
                    break

            # Split the table vertically into two separate tables
            table1 = worksheet.iter_rows(
                min_row=2, max_row=separator_index, values_only=True)
            table2 = worksheet.iter_rows(
                min_row=separator_index+3, values_only=True)

            # Extract data from table 1
            atom_col_idx = 0
            n_col_idx = 1
            li_col_idx = 2
            tot_di_col_idx = 3

            atoms_names = []
            total_electron = []
            loc_electron = []
            deloc_electron = []

            for row in table1:
                atoms_names.append(row[atom_col_idx])
                total_electron.append(float(row[n_col_idx]))
                loc_electron.append(float(row[li_col_idx]))
                deloc_electron.append(float(row[tot_di_col_idx]))

            # print(atoms_names, total_electron, loc_electron, deloc_electron)

            number_of_atoms = len(atoms_names)

            # create empty lists of equal length to atoms_names
            molecule_name = [''] * (len(atoms_names))
            intra_molecule = [0.0] * len(atoms_names)
            inter_molecule = [0.0] * len(atoms_names)

            # Extract data from table 2
            atoms_pairs = []
            total_deloc_electron = []
            A_contri = []
            B_contri = []

            for row in table2:
                atoms_pairs.append(row[atom_col_idx])
                total_deloc_electron.append(float(row[n_col_idx]))
                A_contri.append(float(row[li_col_idx]))
                B_contri.append(float(row[tot_di_col_idx]))

            # print(atoms_pairs, total_deloc_electron, A_contri, B_contri)

            print("Parsing SUM file...")
            print("SUM file has been parsed successfully...")

            # delete the temporary excel file
            # os.remove('example.xlsx')

            print("Done using openpyxl...")

            _ = {
                "number_of_atoms": number_of_atoms,
                "atoms_names": atoms_names,
                "total_electron": total_electron,
                "loc_electron": loc_electron,
                "deloc_electron": deloc_electron,
                "molecule_name": molecule_name,
                "intra_molecule": intra_molecule,
                "inter_molecule": inter_molecule,
                "atoms_pairs": atoms_pairs,
                "total_deloc_electron": total_deloc_electron,
                "A_contri": A_contri,
                "B_contri": B_contri,
                "image_filename": None,
            }

            # print(_)

            self.container.data_for_this_file = _

        except Exception as e:
            msg = CTkMessagebox(
                title="Warning!!", message="Error parsing SUM file\n"+str(e),
                icon="warning", master=app)
            if msg.get() == "OK":
                pass
            return


        text_to_display = ""

        text_to_display += "No of atoms: \t" + str(number_of_atoms) + "\n"
        text_to_display += "Atom names: \t"
        for idx, atom_name in enumerate(atoms_names):
            text_to_display += atom_name + ', '
            # if idx % 8 == 0 and idx != 0:
            #     text_to_display += "\n\t\t"
        # text_to_display += "\n\n"

        self.container.input_frame.info_label.configure(state="normal")
        self.container.input_frame.info_label.delete("1.0", "end")
        self.container.input_frame.info_label.insert("end", text_to_display)
        self.container.input_frame.info_label.configure(state="disabled")

        self.container.input_frame.from_window = "home"
        
        
        # self.container.input_frame.image_filename_entry.configure( state="disabled")

        self.container.show_frame(self.container.input_frame)
        # msg = CTkMessagebox(
        #     title="Info about the extracted data!", message=text_to_display, icon=None, width=500, button_width=120, button_height=30, widen=1)
        # if msg.get() == "OK":
        #     pass

    def get_str_from_numbers_list(self, numbers):
        ranges = []
        start = end = numbers[0]

        for i in range(1, len(numbers)):
            if numbers[i] == end + 1:
                end = numbers[i]
            else:
                if start == end:
                    ranges.append(str(start))
                else:
                    ranges.append(str(start) + '-' + str(end))
                start = end = numbers[i]

        # append the last range
        if start == end:
            ranges.append(str(start))
        else:
            ranges.append(str(start) + '-' + str(end))

        return ','.join(ranges)


class ATOMIC_InputFrame(customtkinter.CTkFrame):
    def __init__(self, master, container, **kwargs):

        super().__init__(master, **kwargs)

        self.container = container

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure(7, weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=9, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="MOWeD-LAC", font=customtkinter.CTkFont(family="Candara",size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.appearance_mode_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=[
                                                                       "Light", "Dark", "System"], command=self.container.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="GUI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%", "130%", "140%", "150%"],
                                                               command=self.container.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # Default value

        self.appearance_mode_optionemenu.set(
            customtkinter.get_appearance_mode())
        self.scaling_optionemenu.set("100%")

        self.mole_system_entry = customtkinter.CTkEntry(
            self, placeholder_text="Enter molecule system name")
        self.mole_system_entry.grid(row=0, column=1, columnspan=3, padx=(
            20, 20), pady=(20, 10), sticky="nsew")

        self.num_mols_label = customtkinter.CTkLabel(
            self, text="Enter number of molecules:", anchor="w")
        self.num_mols_label.grid(row=1, column=1, columnspan=2, padx=(
            20, 20), pady=(0, 10), sticky="nsew")

        self.num_mols_entry = CustomSpinbox(
            self, width=150, step_size=1, data_type=int)
        self.num_mols_entry.grid(row=1, column=3, padx=(
            20, 20), pady=(0, 10), sticky="nsew")
        self.num_mols_entry.set(2)

        self.mol_input_label = customtkinter.CTkLabel(
            self, text="Enter unique molecule names and atom numbers (one per line):\n(Example: H20-1\t1-3)", anchor="w")
        self.mol_input_label.grid(row=2, column=1, columnspan=3, padx=(
            20, 20), pady=(10, 0), sticky="nsew")

        self.mol_input = customtkinter.CTkTextbox(
            self, corner_radius=10, border_width=1)
        self.mol_input.grid(row=3, column=1, columnspan=3,
                            padx=(20, 20), pady=(10, 0), sticky="nsew")
        self.mol_input.configure(text_color="gray")
        self.mol_input.insert(END, "H20-1\t1-3 \nH20-2\t4-6\n\n\n\n")

        # Bind the functions to the focus events of the widget
        self.mol_input.bind('<FocusIn>', self.on_click)
        self.mol_input.bind('<FocusOut>', self.on_leave)

        # self.frag_inpu2 = customtkinter.CTkTextbox(self, corner_radius=10, border_width=1)
        # self.frag_inpu2.grid(row=2, column=2, padx=(20,20), pady=(10,0), sticky="nsew")
        self.cov_threshold_label = customtkinter.CTkLabel(
            self, text="Covalent bond minimum electrons shared threshold:", anchor="w")
        self.cov_threshold_label.grid(
            row=4, column=1, columnspan=2, padx=(20, 0), pady=(10, 0), sticky="nsew")

        self.cov_threshold_entry = CustomSpinbox(
            self, width=150, step_size=0.1, data_type=float)
        self.cov_threshold_entry.grid(row=4, column=3, padx=(
            20, 20), pady=(10, 0), sticky="nsew")
        self.cov_threshold_entry.set(1.2)

        self.image_filename_entry = customtkinter.CTkEntry(
            self, placeholder_text="Enter the image filename")
        self.image_filename_entry.grid(
            row=5, column=1, columnspan=2, padx=(20, 0), pady=(20, 10), sticky="nsew")

        self.browse_image_button = customtkinter.CTkButton(
            self, text="Browse",  command=lambda: self.handle_browse_image_button(self.image_filename_entry))
        self.browse_image_button.grid(row=5, column=3, padx=(
            10, 20), pady=(20, 10), sticky="nsew")

        self.info_label = customtkinter.CTkTextbox(
            self, height=80,fg_color="transparent")
        
        self.info_label.grid(row=6, column=1, columnspan=3, padx=(
            20, 20), pady=(0, 0), sticky="nsew")
        



        self.left_bottom_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.left_bottom_frame.grid(
            row=8, column=1, padx=20, pady=(0, 20), sticky="nsew")

        self.help_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2753", command=self.open_help, border_width=0, width=30,height=29)
        self.help_button.grid(row=0, column=0, padx=(0,10),
                                pady=0)



        self.back_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2190", command=lambda: self.back_to_home(), width=30, height=30)
        self.back_button.grid(row=0, column=1, padx=(10,0),
                              pady=0, sticky="w")

        self.submit_button = customtkinter.CTkButton(
            self, text="Submit", command=lambda: self.handle_submit_button_click())
        self.submit_button.grid(row=8, column=3, padx=(
            20, 20), pady=(0, 20), sticky="nsew")

    def open_help(self):
        self.container.open_help(2)

    def back_to_home(self):
        if self.from_window == "home":
            self.container.show_frame(self.container.home_frame)
        else:
            self.container.show_frame(self.container.page2_frame)
        # self.container.show_frame(self.container.home_frame)

    def on_click(self, event):
        """
        This function is called when the Text widget is clicked/focused.
        It removes the default text if it's still there.
        """
        if self.mol_input.get(1.0, "end-1c") == "H20-1\t1-3 \nH20-2\t4-6\n\n\n\n":
            self.mol_input.delete(1.0, END)
            self.mol_input.configure(text_color=("black", "white"))

    def on_leave(self, event):
        """
        This function is called when the Text widget loses focus.
        It adds back the default text if no text was entered.
        """
        if not self.mol_input.get(1.0, "end-1c"):
            self.mol_input.insert(1.0, "H20-1\t1-3 \nH20-2\t4-6\n\n\n\n")
            self.mol_input.configure(text_color='grey')

    def handle_browse_image_button(self, entry: customtkinter.CTkEntry):

        image_filename = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            title="Select an image",
            filetypes=[("PNG files", "*.png"), ("JPEG files", "*.jpg")]
        )

        if image_filename:
            try:

                self.container.data_for_this_file_image = image_filename
                file_dir, file_name = os.path.split(image_filename)
                print("Image filename: ", file_name)
                os.startfile(image_filename)
                entry.delete(0, "end")
                entry.insert(0, file_name)

            except:
                msg = CTkMessagebox(
                    title="Error", message="An error occured while loading the image!", icon="cancel", master=app)
                if msg.get() == "OK":
                    pass
                return

        else:
            if not self.container.data_for_this_file_image:
                msg = CTkMessagebox(
                    title="Error", message="Please select an image!", icon="cancel", master=app)
                if msg.get() == "OK":
                    pass
            return

    def get_atom_pair(self, string):
        """Extracts two atoms from a string formatted like 'x_y'."""
        string = string.split('_')
        return string[0], string[1]

    def get_atom_number(self, string, atoms_names):
        """Finds the index of an atom in the list of atom names."""
        for i in range(len(atoms_names)):
            if atoms_names[i] == string:
                return i+1

    def is_intra_molecule(self, pair, molecule_relations):
        """Returns True if the given pair of atoms are in the same molecule."""
        atom_names = self.container.data_for_this_file["atoms_names"]
        A, B = self.get_atom_pair(pair)
        A_num, B_num = self.get_atom_number(
            A, atom_names), self.get_atom_number(B, atom_names)
        return molecule_relations[A_num] == molecule_relations[B_num]

    def next_cell(self, cell):
        """Returns the next cell to the right of the given cell."""
        return cell.offset(row=0, column=1)

    def handle_submit_button_click(self):

        _ = self.container.data_for_this_file

        number_of_atoms = _["number_of_atoms"]
        atoms_names = _["atoms_names"]
        total_electron = _["total_electron"]
        loc_electron = _["loc_electron"]
        deloc_electron = _["deloc_electron"]
        molecule_name = _["molecule_name"]
        intra_molecule = _["intra_molecule"]
        inter_molecule = _["inter_molecule"]
        atoms_pairs = _["atoms_pairs"]
        total_deloc_electron = _["total_deloc_electron"]
        A_contri = _["A_contri"]
        B_contri = _["B_contri"]
        image_filename = self.container.data_for_this_file_image
        print("Printing image filename: ", image_filename)

        format_error_message = "Please enter molecule names and atom numbers in form \"XXXX XX-XX\"."

        try:
            number_of_molecules = int(self.num_mols_entry.get())

            cov_threshold = float(self.cov_threshold_entry.get())

            if cov_threshold < 0 or number_of_molecules < 0:
                CTkMessagebox(
                    title="Error", message="Please enter positive values!", icon="cancel", master=app)
                return

        except:
            return

        if self.mole_system_entry.get() == "":
            CTkMessagebox(
                title="Error", message="Please enter the molecule system!", icon="cancel", master=app)
            return

        if self.num_mols_entry.get() == "":
            CTkMessagebox(
                title="Error", message="Please enter the number of molecules!", icon="cancel", master=app)
            return

        if self.mol_input.get("1.0", "end-1c") == "" or self.mol_input.get("1.0", "end-1c") == "H20-1\t1-3 \nH20-2\t4-6\n\n\n\n":
            CTkMessagebox(
                title="Error", message="Please enter the molecules!", icon="cancel", master=app)
            return

        if self.cov_threshold_entry.get() == "":
            CTkMessagebox(
                title="Error", message="Please enter the coverage threshold!", icon="cancel", master=app)
            return

        try:

            doing_msg = ProgressCTkMessagebox(
                title="Please Wait...", message="Organizing data...",
                icon="info", fade_in_duration=1, master=app)

            doing_msg.set_progress(0)
            print("setting progress to 0")

            print("cov_threshold: " + str(cov_threshold))

            all_molecules = [line for line in self.mol_input.get(
                "1.0", "end-1c").split("\n") if line]
            system_name = self.mole_system_entry.get()

            if len(all_molecules) != number_of_molecules:
                doing_msg.destroy()
                CTkMessagebox(
                    title="Error", message="Number of molecules entered does not match number of molecules specified.", icon="cancel", master=app)
                return

            molecule_names = []
            molecules_numbers = []
            molecule_relations = {}
            try:
                for i in range(number_of_molecules):
                    temp = all_molecules[i]
                    temp = temp.replace("\t", " ")
                    temp = temp.split()
                    if len(temp) != 2:
                        doing_msg.destroy()
                        CTkMessagebox(
                            title="Error", message=format_error_message, icon="cancel", master=app)
                        return
                    molecule_names.append(temp[0])
                    numbers = self.get_numbers(temp[1])
                    molecules_numbers.append(numbers)

                    for i in numbers:
                        molecule_relations[i] = temp[0]
            except IndexError:
                doing_msg.destroy()
                CTkMessagebox(
                    title="Error", message=format_error_message, icon="cancel", master=app)
                return

            if sum([len(molecules_numbers[i]) for i in range(number_of_molecules)]) != number_of_atoms:
                doing_msg.destroy()
                CTkMessagebox(
                    title="Error", message="Number of atoms entered does not match number of atoms detected.", icon="cancel", master=app)
                return

            all_atoms_check = []
            for i in range(number_of_molecules):
                for j in range(len(molecules_numbers[i])):
                    all_atoms_check.append(molecules_numbers[i][j])
            all_atoms_check.sort()
            for i in range(number_of_atoms):
                if all_atoms_check[i] != i+1:
                    doing_msg.destroy()
                    CTkMessagebox(
                        title="Error", message="Atom numbers entered are not correct.", icon="cancel", master=app)
                    return
            # check if molecule names are unique
            if len(molecule_names) != len(set(molecule_names)):
                doing_msg.destroy()
                CTkMessagebox(
                    title="Error", message="Molecule names are not unique.", icon="cancel", master=app)
                return

            # you can do whatever you want with the data here (e.g. save it to a file,
            # display it in a new window, etc.)
            print("Number of molecules:", number_of_molecules)
            print("Molecule names:")
            for name in molecule_names:
                print("- " + name)

            print("Molecule numbers:")
            for number in molecules_numbers:
                print("- " + str(number))

            print("Molecule relations:")
            for key, value in molecule_relations.items():
                print("- " + str(key) + ": " + value)

            for i in range(1, len(molecule_relations)+1):
                # print(i)
                if i == 1:
                    molecule_name[i-1] = molecule_relations[i]
                elif molecule_relations[i] == molecule_relations[i-1]:
                    molecule_name[i-1] = ''
                else:
                    molecule_name[i-1] = molecule_relations[i]

            # loop through each atom name and add it as a key to the data_of_table2 dictionary
            data_of_table2 = {}
            reversed_data_of_table2 = {}
            sorted_data_of_table2 = {}
            reversed_sorted_data_of_table2 = {}
            reversed_data_of_table2_inter = {}

            reversed_data_of_table2_intra = {}
            reversed_data_of_table2_intra_cov = {}

            sorted_atoms_pair_total_deloc_electron = {}

            sorted_according_to_contri = {}

            intra_mol_data = {}
            inter_mol_data = {}

            intra_mol_dict = {}
            inter_mol_dict = {}

            intra_sorted_according_to_total_deloc_electron = {
                "covalent": {},
                "non-covalent": {}
            }

            intra_sorted_according_to_contri = {
                "covalent": {},
                "non-covalent": {}
            }

            data_of_table2_intra = {
                "covalent": {},
                "non-covalent": {}
            }

            sorted_data_of_table2_intra = {
                "covalent": {},
                "non-covalent": {}
            }

            data_of_table2_inter = {}

            sorted_data_of_table2_inter = {}

            inter_sorted_according_to_total_deloc_electron = {}

            inter_sorted_according_to_contri = {}

            for atom_name in atoms_names:
                data_of_table2[atom_name] = {}
                reversed_data_of_table2[atom_name] = {}

                reversed_data_of_table2_intra[atom_name] = {}
                reversed_data_of_table2_inter[atom_name] = {}
                reversed_data_of_table2_intra_cov[atom_name] = {}

                data_of_table2_intra["covalent"][atom_name] = {}
                data_of_table2_intra["non-covalent"][atom_name] = {}

                data_of_table2_inter[atom_name] = {}

            for name in molecule_names:
                intra_mol_dict[name] = {}
                intra_mol_dict[name]["covalent"] = {}
                intra_mol_dict[name]["non-covalent"] = {}

                inter_mol_dict[name] = {}

            # loop through each set of atom pairs and associated values
            for atom_pair, total_deloc, A, B in zip(atoms_pairs, total_deloc_electron, A_contri, B_contri):
                # get the names and numbers of the two atoms in the pair
                # print(atom_pair)
                A_name, B_name = self.get_atom_pair(atom_pair)
                A_index = self.get_atom_number(A_name, atoms_names)
                B_index = self.get_atom_number(B_name, atoms_names)

                a_to_b = A_name + "_" + B_name
                b_to_a = B_name + "_" + A_name
                sorted_according_to_contri[a_to_b] = A
                sorted_according_to_contri[b_to_a] = B

                # check whether the atoms belong to the same molecule or not, and update intra- or inter-molecule values accordingly
                if molecule_relations[A_index] == molecule_relations[B_index]:
                    intra_molecule[A_index-1] += A
                    intra_molecule[B_index-1] += B
                else:
                    inter_molecule[A_index-1] += A
                    inter_molecule[B_index-1] += B

                # add values to the data_of_table2 dictionary
                data_of_table2[A_name][B_name] = A
                data_of_table2[B_name][A_name] = B

            atom_molecule_dict = {}

            for atom_name in atoms_names:
                for idx, numbers in enumerate(molecules_numbers):
                    atom_molecule_dict[atom_name+"_Molecule-" +
                                       str(idx+1)] = [0.0, 0.0, 0.0]
                    for num in numbers:
                        if atom_name != atoms_names[num-1]:
                            atom_molecule_dict[atom_name+"_Molecule-" + str(
                                idx+1)][1] += data_of_table2[atom_name][atoms_names[num-1]]
                            atom_molecule_dict[atom_name+"_Molecule-" + str(
                                idx+1)][2] += data_of_table2[atoms_names[num-1]][atom_name]

                    atom_molecule_dict[atom_name+"_Molecule-" + str(idx+1)][0] = atom_molecule_dict[atom_name+"_Molecule-" + str(
                        idx+1)][1] + atom_molecule_dict[atom_name+"_Molecule-" + str(idx+1)][2]
            # if number_of_molecules > 1:
                atom_molecule_dict[atom_name+"_System"] = [0.0, 0.0, 0.0]
                for idx, numbers in enumerate(molecules_numbers):
                    atom_molecule_dict[atom_name +
                                    "_System"][0] += atom_molecule_dict[atom_name+"_Molecule-" + str(idx+1)][0]
                    atom_molecule_dict[atom_name +
                                    "_System"][1] += atom_molecule_dict[atom_name+"_Molecule-" + str(idx+1)][1]
                    atom_molecule_dict[atom_name +
                                    "_System"][2] += atom_molecule_dict[atom_name+"_Molecule-" + str(idx+1)][2]

            for key, value in data_of_table2.items():
                for k, v in value.items():
                    reversed_data_of_table2[k][key] = v

            for atom_name in atoms_names:
                sorted_data_of_table2[atom_name] = dict(sorted(
                    data_of_table2[atom_name].items(), key=lambda item: item[1], reverse=True))

            for atom_name in atoms_names:
                reversed_sorted_data_of_table2[atom_name] = dict(sorted(
                    reversed_data_of_table2[atom_name].items(), key=lambda item: item[1], reverse=True))

            # sort various data structures by electron values and store in a dictionary
            sorted_data = {
                'N(A)': dict(sorted(dict(zip(atoms_names, total_electron)).items(), key=lambda item: item[1], reverse=True)),
                'loc-N(A)': dict(sorted(dict(zip(atoms_names, loc_electron)).items(), key=lambda item: item[1], reverse=True)),
                'total': dict(sorted(dict(zip(atoms_names, deloc_electron)).items(), key=lambda item: item[1], reverse=True)),
                'intra-Mol': dict(sorted(dict(zip(atoms_names, intra_molecule)).items(), key=lambda item: item[1], reverse=True)),
                'inter-Mol': dict(sorted(dict(zip(atoms_names, inter_molecule)).items(), key=lambda item: item[1], reverse=True)),
            }

            sorted_atoms_pair_total_deloc_electron = dict(sorted(dict(zip(
                atoms_pairs, total_deloc_electron)).items(), key=lambda item: item[1], reverse=True))

            sorted_according_to_contri = dict(sorted(
                sorted_according_to_contri.items(), key=lambda item: item[1], reverse=True))

            for idx, values in enumerate(zip(atoms_pairs, total_deloc_electron, A_contri, B_contri)):

                if self.is_intra_molecule(values[0], molecule_relations):
                    intra_mol_data[values[0]] = [
                        values[1], values[2], values[3]]
                else:
                    inter_mol_data[values[0]] = [
                        values[1], values[2], values[3]]

            print("cov_threshold: " + str(cov_threshold))

            for keys, values in intra_mol_data.items():

                if values[0] > cov_threshold:
                    intra_mol_dict[molecule_relations[self.get_atom_number(
                        self.get_atom_pair(keys)[0], atoms_names)]]["covalent"][keys] = values

                else:
                    intra_mol_dict[molecule_relations[self.get_atom_number(
                        self.get_atom_pair(keys)[0], atoms_names)]]["non-covalent"][keys] = values
            # print("intra_mol_dict: " + str(intra_mol_dict))
            print()

            inter_mol_dict_pair = {}
            for keys, values in inter_mol_data.items():
                mol1 = molecule_relations[self.get_atom_number(
                    self.get_atom_pair(keys)[0], atoms_names)]
                mol2 = molecule_relations[self.get_atom_number(
                    self.get_atom_pair(keys)[1], atoms_names)]
                inter_mol_dict[mol1][keys] = values

                if mol1+"_"+mol2 not in inter_mol_dict_pair.keys():
                    inter_mol_dict_pair[mol1+"_"+mol2] = {}
                inter_mol_dict_pair[mol1+"_"+mol2][keys] = values

            # print("intra_mol_dict: " + str(inter_mol_dict_pair))

            for name in molecule_names:

                for keys, values in intra_mol_dict[name]["covalent"].items():
                    intra_sorted_according_to_total_deloc_electron["covalent"][keys] = values[0]

                    A, B = self.get_atom_pair(keys)
                    a_to_b = A + "_" + B
                    b_to_a = B + "_" + A

                    intra_sorted_according_to_contri["covalent"][a_to_b] = values[1]
                    intra_sorted_according_to_contri["covalent"][b_to_a] = values[2]

                for keys, values in intra_mol_dict[name]["non-covalent"].items():
                    intra_sorted_according_to_total_deloc_electron["non-covalent"][keys] = values[0]

                    A, B = self.get_atom_pair(keys)
                    a_to_b = A + "_" + B
                    b_to_a = B + "_" + A

                    intra_sorted_according_to_contri["non-covalent"][a_to_b] = values[1]
                    intra_sorted_according_to_contri["non-covalent"][b_to_a] = values[2]

                for keys, values in inter_mol_dict[name].items():
                    inter_sorted_according_to_total_deloc_electron[keys] = values[0]

                    A, B = self.get_atom_pair(keys)
                    a_to_b = A + "_" + B
                    b_to_a = B + "_" + A

                    inter_sorted_according_to_contri[a_to_b] = values[1]
                    inter_sorted_according_to_contri[b_to_a] = values[2]

            intra_sorted_according_to_total_deloc_electron["covalent"] = dict(sorted(
                intra_sorted_according_to_total_deloc_electron["covalent"].items(), key=lambda item: item[1], reverse=True))

            intra_sorted_according_to_total_deloc_electron["non-covalent"] = dict(sorted(
                intra_sorted_according_to_total_deloc_electron["non-covalent"].items(), key=lambda item: item[1], reverse=True))

            inter_sorted_according_to_total_deloc_electron = dict(sorted(
                inter_sorted_according_to_total_deloc_electron.items(), key=lambda item: item[1], reverse=True))

            inter_sorted_according_to_contri = dict(sorted(
                inter_sorted_according_to_contri.items(), key=lambda item: item[1], reverse=True))

            intra_sorted_according_to_contri["covalent"] = dict(sorted(
                intra_sorted_according_to_contri["covalent"].items(), key=lambda item: item[1], reverse=True))

            intra_sorted_according_to_contri["non-covalent"] = dict(sorted(
                intra_sorted_according_to_contri["non-covalent"].items(), key=lambda item: item[1], reverse=True))

            for name in molecule_names:
                covalent_dict = intra_mol_dict[name]["covalent"]
                non_covalent_dict = intra_mol_dict[name]["non-covalent"]
                inter_this_mol_dict = inter_mol_dict[name]

                for keys, values in covalent_dict.items():
                    A, B = self.get_atom_pair(keys)
                    data_of_table2_intra["covalent"][A][B] = values[1]
                    data_of_table2_intra["covalent"][B][A] = values[2]
                for keys, values in non_covalent_dict.items():
                    A, B = self.get_atom_pair(keys)
                    data_of_table2_intra["non-covalent"][A][B] = values[1]
                    data_of_table2_intra["non-covalent"][B][A] = values[2]
                for keys, values in inter_this_mol_dict.items():
                    A, B = self.get_atom_pair(keys)
                    data_of_table2_inter[A][B] = values[1]
                    data_of_table2_inter[B][A] = values[2]

            # now sort this data individually

            for keys, values in data_of_table2_intra["covalent"].items():
                sorted_data_of_table2_intra["covalent"][keys] = dict(
                    sorted(values.items(), key=lambda item: item[1], reverse=True))

            for keys, values in data_of_table2_intra["non-covalent"].items():
                sorted_data_of_table2_intra["non-covalent"][keys] = dict(
                    sorted(values.items(), key=lambda item: item[1], reverse=True))

            for keys, values in data_of_table2_inter.items():
                sorted_data_of_table2_inter[keys] = dict(
                    sorted(values.items(), key=lambda item: item[1], reverse=True))
        except Exception as e:
            doing_msg.destroy()
            msg = CTkMessagebox(
                title="Error", message="Error while extracting data from the file. Please check the file and try again.\n"+str(e), icon="cancel", master=app)

        try:
            doing_msg.set_progress(
                0.15, text_="Working on atoms sheet", speed=60)

            # create a new excel workbook and select the active worksheet
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'atoms'

            # define cell styles
            header_fill = PatternFill(
                start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            header_font = Font(bold=True)
            total_fill = PatternFill(start_color='DCE6F1',
                                     end_color='DCE6F1', fill_type='solid')
            total_border = Border(
                top=Side(border_style='thin', color='FF000000'))
            total_font = openpyxl.styles.Font(bold=True)

            # write the header row
            headers = ['Atom A',
                       'N(A)', 'loc-N(A)', 'total', 'intra-Mol', 'inter-Mol']
            for col_idx, header in enumerate(headers):
                cell = worksheet.cell(row=1, column=col_idx+2)

                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                if col_idx >= 1:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

            # write the data rows
            molu_cnt = 1
            for row_idx, values in enumerate(zip(molecule_name, atoms_names, total_electron, loc_electron, deloc_electron, intra_molecule, inter_molecule), start=2):
                for col_idx, value in enumerate(values):
                    cell = worksheet.cell(row=row_idx, column=col_idx+1)
                    if col_idx == 0 and value != "":
                        cell.value = "Molecule-"+str(molu_cnt)
                        cell.font = total_font
                        molu_cnt += 1
                    else:
                        cell.value = value
                    cell.number_format = '0.0000'

            # compute and write totals for each column
            col_idx_start = 7
            for key, value_dict in sorted_data.items():
                col_idx_start += 3
                cell = worksheet.cell(row=1, column=col_idx_start)
                cell.value = "Atom A"
                cell.fill = header_fill
                cell.font = header_font

                cell = self.next_cell(cell)
                cell.value = key
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")

                for row_idx, values in enumerate(value_dict.items(), start=2):
                    for col_idx, value in enumerate(values):
                        cell = worksheet.cell(
                            row=row_idx, column=col_idx+col_idx_start)
                        cell.value = value
                        cell.number_format = '0.0000'

                cell = cell.offset(row=1, column=-1)
                cell.value = 'Total:'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = sum(value_dict.values())
                cell.number_format = '0.0000'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

            # compute and write totals for the last row of data
            cell = worksheet.cell(row=len(atoms_names)+2, column=2)
            cell.value = 'Total:'
            cell.border = total_border
            cell.font = total_font
            cell.fill = total_fill

            cell = self.next_cell(cell)
            total_values = [
                sum(total_electron),
                sum(loc_electron),
                sum(deloc_electron),
                sum(intra_molecule),
                sum(inter_molecule),
            ]
            total_for_each_molecule = {}

            for idx, data in enumerate(zip(molecule_names, molecules_numbers)):
                name = data[0]
                numbers = data[1]

                total_for_each_molecule[name] = [sum([total_electron[i-1] for i in numbers]),
                                                 sum([loc_electron[i-1]
                                                      for i in numbers]),
                                                 sum([deloc_electron[i-1]
                                                      for i in numbers]),
                                                 sum([intra_molecule[i-1]
                                                      for i in numbers]),
                                                 sum([inter_molecule[i-1] for i in numbers])]

            for value in total_values:
                cell.value = value
                cell.number_format = '0.0000'
                cell.border = total_border
                cell.font = total_font
                cell.fill = total_fill
                cell = self.next_cell(cell)

            total_properties = {
                'fill': total_fill,
                'font': total_font,
            }
            molu_cnt = 1
            cell = cell.offset(row=2, column=-6)
            if number_of_molecules > 1:
                add_cell_value(worksheet, cell.row, cell.column,
                               'Sub-Total:', **total_properties)
            cell = cell.offset(row=1)
            for key, values in total_for_each_molecule.items():
                if number_of_molecules > 1:
                    cell.value = "Molecule-"+str(molu_cnt)+":"
                    # cell.border = total_border
                    cell.font = total_font
                    cell.fill = total_fill
                # cell = self.next_cell(cell)
                cell = self.next_cell(cell)

                for value in values:
                    if number_of_molecules > 1:

                        cell.value = value
                        cell.number_format = '0.0000'
                        # cell.border = total_border
                        cell.font = total_font
                        cell.fill = total_fill
                    cell = self.next_cell(cell)

                cell = cell.offset(column=2)
                cell.value = "Molecule-"+str(molu_cnt)+":"
                molu_cnt += 1
                cell.font = total_font
                cell = self.next_cell(cell)
                cell = self.next_cell(cell)

                cell.value = key

                cell = cell.offset(row=1, column=-10)

            worksheet.insert_rows(1, 6)
            # worksheet.insert_cols(1, 1)

            # Define common style properties
            heading_style = NamedStyle(name="heading")
            heading_style.font = Font(bold=True)
            heading_style.alignment = Alignment(
                horizontal="center", vertical="center")
            heading_style.border = Border(left=Side(style="thick"), right=Side(
                style="thick"), top=Side(style="thick"), bottom=Side(style="thick"))

            # Set values and styles for each cell
            worksheet.merge_cells("B2:W3")
            worksheet["B2"].value = "The total electron population found, on average, in a space occupied by an atom A, atom-N(A). The atom-N(A) count is partitioned to electrons (i) found only in atom A (the loc-N(A) electron count), (ii) delocalized to atoms of the same molecule (the intra-Mol electron count) and (iii) delocalized to other molecules (the inter-Mol electron count). These values are computed for all atoms of a molecular system."
            worksheet["B2"].style = heading_style



            worksheet.merge_cells("B5:G5")
            worksheet["B5"].value = "Atomic electron counts, atom-N(A)"
            worksheet["B5"].style = heading_style

            worksheet.merge_cells("J6:N6")
            worksheet["J6"].value = "Total and localized electron counts"
            worksheet["J6"].style = heading_style

            worksheet.merge_cells("J5:N5")
            worksheet["J5"].value = "Electron counts are printed in the descending order"
            worksheet["J5"].fill = total_fill
            worksheet["J5"].font = Font(bold=True, color="6f30a0")
            worksheet["J5"].alignment = Alignment(
                horizontal="center", vertical="center")

            worksheet.merge_cells("P6:W6")
            worksheet["P6"].value = "Total deloc-N(A), intramolecular and intermolecular atomic deloc-N(A) counts"
            worksheet["P6"].style = heading_style
            worksheet.merge_cells("P5:W5")
            worksheet["P5"].value = "Delocalized electron counts are printed in the descending order"
            worksheet["P5"].fill = total_fill
            worksheet["P5"].font = Font(bold=True, color="6f30a0")
            worksheet["P5"].alignment = Alignment(
                horizontal="center", vertical="center")


            # Set border style for merged ranges
            for range_string in ["B5:G5", "J6:N6", "P6:W6","B2:W3"]:
                cells_range = worksheet[range_string]
                for row in cells_range:
                    for cell in row:
                        cell.style = heading_style

            worksheet["B2"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

            worksheet.merge_cells("E6:G6")
            cell = worksheet.cell(row=6, column=5)
            cell.value = "deloc-N(A)"
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(
                horizontal='center', vertical='center')
            cell.fill = header_fill

            def insert_image(image_path, worksheet, coordinates):
                try:
                    # Load image and calculate new dimensions to match required height
                    img = Image(image_path)
                    width, height = img.width, img.height
                    aspect_ratio = width / height
                    new_height = 377.946666667  # 10 cm in pixels (at 96 dpi)
                    new_width = new_height * aspect_ratio
                    img.height = new_height
                    img.width = new_width

                    # Insert the image into the specified cell
                    cell = worksheet[coordinates]
                    worksheet.add_image(img, coordinates)
                except:
                    print("Failed to insert image: {}".format(image_path))

            # Call the function for both images with their respective coordinates
            default_image_filename = os.path.join(
                base_path, 'assets', 'MOWeD-LAC.png')

            insert_image(default_image_filename, worksheet,
                         'C'+str(len(atoms_names)+16+number_of_molecules))
            if image_filename == "" or image_filename == None:
                print("No Image selected")
            else:
                insert_image(image_filename, worksheet,
                             'K'+str(len(atoms_names)+16+number_of_molecules))

            mole_name_properties = {
                'font': Font(bold=True, color='FF0000'),
            }

            # worksheet.merge_cells("K2:M2")
            # add_cell_value(worksheet, 2, 11,
            #                "In the descending order", **mole_name_properties)
            worksheet.merge_cells("B"+str(len(atoms_names)+13+number_of_molecules) +
                                  ":G"+str(len(atoms_names)+13+number_of_molecules))
            worksheet["B"+str(len(atoms_names)+13+number_of_molecules)
                      ].value = "Partitioning of the total atomic electron population, N(A)"
            worksheet["B"+str(len(atoms_names)+13 +
                              number_of_molecules)].font = Font(bold=True)

            worksheet.merge_cells("J"+str(len(atoms_names)+13+number_of_molecules) +
                                  ":L"+str(len(atoms_names)+13+number_of_molecules))
            worksheet["J"+str(len(atoms_names)+13+number_of_molecules)
                      ].value = "Molecular system:"
            worksheet["M"+str(len(atoms_names)+13 +
                              number_of_molecules)].value = system_name
            worksheet["J"+str(len(atoms_names)+13 +
                              number_of_molecules)].font = Font(bold=True)

            worksheet.insert_rows(1, 1)
            worksheet.insert_cols(1, 1)

            for merged_cell in worksheet.merged_cells:
                merged_cell.shift(1, 1)

            # # set the width of the columns
            # widths_data = {
            #     "C": 15,

            # }

            # change the width of all the columns
            for col in worksheet.columns:
                key = col[0].column_letter
                worksheet.column_dimensions[key].width = 11.78

            # save the workbook
            # workbook.save('table1_.xlsx')
            # create a new excel workbook and select the active worksheet
            # workbook2 = openpyxl.Workbook()
            # worksheet2 = workbook2.active

            doing_msg.set_progress(
                0.25, text_="Working on atom-pairs(A,B) sheet", speed=50)

            worksheet2 = workbook.create_sheet("atom-pairs(A,B)")

            # define cell styles
            header_fill = PatternFill(
                start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            # write the header row
            headers = ["(A,B) count", "Atom A",
                       "Atom B", "(A,B)", "A to B", "B to A"]
            for col_idx, header in enumerate(headers):
                cell = worksheet2.cell(row=2, column=col_idx+1)
                cell.value = header
                cell.fill = header_fill
                cell.font = openpyxl.styles.Font(bold=True)
                if col_idx > 2:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center', vertical='center')
            # # Write the data rows
            # for row in range(len(atoms_names)):
            #     worksheet2.append(["",atoms_names[row], total_electron[row], loc_electron[row], deloc_electron[row]])
            # write the data rows

            # create lists for identifying atoms A and B in each atom pair
            Atoms_A, Atoms_B = zip(*(atom_pair.split('_')
                                     for atom_pair in atoms_pairs))

            # write data to worksheet2 table
            for row_idx, values in enumerate(zip(range(1, len(Atoms_A)+10), Atoms_A, Atoms_B, total_deloc_electron, A_contri, B_contri), start=3):
                for col_idx, value in enumerate(values):
                    cell = worksheet2.cell(row=row_idx, column=col_idx+1)
                    cell.value = value
                    if col_idx != 0:
                        cell.number_format = '0.0000'

            # insert new columns into worksheet2 for each atom name
            for atom_name in reversed(atoms_names):
                worksheet2.insert_cols(9, 4)

                # write header row for each atom
                cell = worksheet2.cell(row=2, column=9)
                cell.value = "To atom B"
                cell.fill = header_fill
                cell.font = header_font
                cell = self.next_cell(cell)
                cell.value = atom_name
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.font = header_font
                cell = cell.offset(row=-1)
                cell.value = "By atom A"
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.font = header_font

                # write data rows for each atom
                for row_idx, values in enumerate(zip(data_of_table2[atom_name].keys(), data_of_table2[atom_name].values()), start=3):
                    for col_idx, value in enumerate(values):
                        cell = worksheet2.cell(row=row_idx, column=col_idx+9)
                        cell.value = value
                        cell.number_format = '0.0000'

                # write footer row for each atom
                cell = cell.offset(row=1, column=-1)
                cell.value = 'Total:'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = sum(data_of_table2[atom_name].values())
                cell.number_format = '0.0000'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                # write secondary header row for each atom
                cell = cell.offset(row=6, column=-1)
                cell.value = 'To atom B'
                cell.fill = header_fill
                cell.font = header_font
                cell = self.next_cell(cell)
                cell.value = atom_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell = cell.offset(row=-1)
                cell.value = 'By atom A'
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell = cell.offset(row=1, column=1)
                cell.value = '''%-fraction'''
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

                # write secondary data rows for each atom
                tot = sum(sorted_data_of_table2[atom_name].values())
                for row_idx, values in enumerate(zip(sorted_data_of_table2[atom_name].keys(), sorted_data_of_table2[atom_name].values()), start=3):
                    for col_idx, value in enumerate(values):
                        cell = worksheet2.cell(
                            row=row_idx+len(atoms_names)+6, column=col_idx+9)
                        cell.value = value
                        cell.number_format = '0.0000'
                        if col_idx == 1:
                            cell = self.next_cell(cell)
                            cell.value = value*100/tot
                            cell.number_format = '0.00'

                # write secondary footer row for each atom
                cell = cell.offset(row=1, column=-2)
                cell.value = 'Total:'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = sum(sorted_data_of_table2[atom_name].values())
                cell.number_format = '0.0000'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = 100
                cell.number_format = '0.00'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                # write header row for each atom
                cell = worksheet2.cell(
                    row=2*number_of_atoms+42+number_of_molecules, column=9)

                cell_row3 = cell.row+3

                cell.value = "By atom B"
                cell.fill = header_fill
                cell.font = header_font
                cell = self.next_cell(cell)
                cell.value = atom_name
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.font = header_font
                cell = cell.offset(row=-1)
                cell.value = "To atom A"
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.font = header_font

                # write data rows for each atom
                for row_idx, values in enumerate(zip(reversed_data_of_table2[atom_name].keys(), reversed_data_of_table2[atom_name].values()), start=2*number_of_atoms+43+number_of_molecules):
                    for col_idx, value in enumerate(values):
                        cell = worksheet2.cell(row=row_idx, column=col_idx+9)
                        cell.value = value
                        cell.number_format = '0.0000'

                # write footer row for each atom
                cell = cell.offset(row=1, column=-1)
                cell.value = 'Total:'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = sum(reversed_data_of_table2[atom_name].values())
                cell.number_format = '0.0000'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                # write secondary header row for each atom
                cell = cell.offset(row=6, column=-1)
                cell_row4 = cell.row+3

                cell.value = 'By atom B'
                cell.fill = header_fill
                cell.font = header_font
                cell = self.next_cell(cell)
                cell.value = atom_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell = cell.offset(row=-1)
                cell.value = 'To atom A'
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell = cell.offset(row=1, column=1)
                cell.value = '''%-fraction'''
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

                tot = sum(reversed_data_of_table2[atom_name].values())
                # write secondary data rows for each atom
                for row_idx, values in enumerate(zip(reversed_sorted_data_of_table2[atom_name].keys(), reversed_sorted_data_of_table2[atom_name].values()), start=cell.row+1):
                    for col_idx, value in enumerate(values):
                        cell = worksheet2.cell(
                            row=row_idx, column=col_idx+9)
                        cell.value = value
                        cell.number_format = '0.0000'
                        if col_idx == 1:
                            cell = self.next_cell(cell)
                            cell.value = value*100/tot
                            cell.number_format = '0.00'

                # write secondary footer row for each atom
                cell = cell.offset(row=1, column=-2)
                cell.value = 'Total:'

                last_cell_row = cell.row
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = sum(
                    reversed_sorted_data_of_table2[atom_name].values())
                cell.number_format = '0.0000'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

                cell = self.next_cell(cell)
                cell.value = 100
                cell.number_format = '0.00'
                cell.fill = total_fill
                cell.font = total_font
                cell.border = total_border

            cell_row5 = last_cell_row+9
            # write header row for each atom
            # headers = ["Atom A", "Molecule#",
            #            "(A,Mol)", "A to Mol", "Mol to A"]

            # for col_idx, header in enumerate(headers):
            #     cell = worksheet2.cell(row=last_cell_row+6, column=col_idx+9)
            #     cell.value = header
            #     cell.fill = header_fill
            #     cell.font = header_font
            #     if col_idx > 1:
            #         cell.alignment = Alignment(
            #             horizontal='center', vertical='center')

            # # write data rows for each atom
            # for row_idx, values in enumerate(atom_molecule_dict.items(), start=last_cell_row+7):
            #     atom, molecule = values[0].split('_')

            #     cell = worksheet2.cell(row=row_idx, column=9)
            #     cell.value = atom
            #     cell = self.next_cell(cell)
            #     cell.value = molecule
            #     cell = self.next_cell(cell)
            #     for col_idx, value in enumerate(values[1]):
            #         cell.value = value
            #         cell.number_format = '0.0000'
            #         cell = self.next_cell(cell)

            worksheet2.insert_cols(7, 9)

            # write the header row
            headers = ["(A,B) count", "Atom A", "Atom B",
                       "(A,B)", "", "(A,B) count", "Atom A", "Atom B", "A to B"]
            for col_idx, header in enumerate(headers):
                if col_idx != 4:
                    cell = worksheet2.cell(row=2, column=col_idx+8)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = openpyxl.styles.Font(bold=True)
                if col_idx == 3 or col_idx == 8:
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center', vertical='center')

            for row_idx, values in enumerate(sorted_atoms_pair_total_deloc_electron.items(), start=3):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-2), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=8):
                    cell = worksheet2.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 8:
                        cell.number_format = '0.0000'

            for row_idx, values in enumerate(sorted_according_to_contri.items(), start=3):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-2), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=13):
                    cell = worksheet2.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 13:
                        cell.number_format = '0.0000'

            # Define the cell properties
            total_properties = {
                'fill': total_fill,
                'font': total_font,
                'border': total_border
            }
            number_properties = {
                'number_format': '0.0000',
                **total_properties  # merge with total_properties
            }

            # Write the footer row
            cell = worksheet2.cell(row=len(atoms_pairs)+3, column=3)
            add_cell_value(worksheet2, cell.row, cell.column,
                           'Total:', **total_properties)

            cell = self.next_cell(cell)
            add_cell_value(worksheet2, cell.row, cell.column, sum(
                total_deloc_electron), **number_properties)

            cell = self.next_cell(cell)
            add_cell_value(worksheet2, cell.row, cell.column,
                           sum(A_contri), **number_properties)

            cell = self.next_cell(cell)
            add_cell_value(worksheet2, cell.row, cell.column,
                           sum(B_contri), **number_properties)

            cell = self.next_cell(self.next_cell(
                self.next_cell(self.next_cell(cell))))
            add_cell_value(worksheet2, cell.row, cell.column,
                           'Total:', **total_properties)

            cell = self.next_cell(cell)
            add_cell_value(worksheet2, cell.row, cell.column, sum(
                sorted_atoms_pair_total_deloc_electron.values()), **number_properties)

            cell = worksheet2.cell(
                row=len(sorted_according_to_contri)+3, column=15)
            add_cell_value(worksheet2, cell.row, cell.column,
                           'Total:', **total_properties)

            cell = self.next_cell(cell)
            add_cell_value(worksheet2, cell.row, cell.column, sum(
                sorted_according_to_contri.values()), **number_properties)

            worksheet2.insert_rows(1, 6)
            worksheet2.insert_cols(1, 2)
            worksheet2.insert_cols(20, 1)
            # worksheet2.insert_cols(1, 1)

            # worksheet2.merge_cells("F7:G7")
            # worksheet2.merge_cells("J6:P6")
            # worksheet2.merge_cells("P7:Q7")
            # worksheet2.merge_cells("S6:Y6")

            heading_properties = {
                'fill': total_fill,
                'font': header_font,
                'alignment': openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            }
            tranparent_heading_properties = {
                "font": header_font,
                "alignment": openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            }
            colored_heading_properties = {
                "fill": total_fill,
                'font': Font(bold=True, color="6f30a0"),
                "alignment": openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
            }

            worksheet2.merge_cells("C3:R4")
            add_cell_value(worksheet2, 3, 3, '''The number of electrons (i) shared by atom-piar (A,B) is shown as 'e-shared', (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A are shown as 'e-delocalized by:'.                                                                                                                         These values are computed for all unique atom-pairs (A,B) of a molecular system''', **tranparent_heading_properties)
            worksheet2.merge_cells("C7:E7")
            add_cell_value(
                worksheet2, 7, 3, 'All unique atom-pairs (A,B)', **tranparent_heading_properties)
            worksheet2.merge_cells("G7:H7")
            add_cell_value(worksheet2, 7, 6, 'e-shared', **
                           tranparent_heading_properties)
            add_cell_value(worksheet2, 7, 7, 'e-delocalized by:',
                           **tranparent_heading_properties)
            worksheet2.merge_cells("J7:M7")
            add_cell_value(
                worksheet2, 7, 10, 'Electons shared by atoms A and B', **tranparent_heading_properties)
            worksheet2.merge_cells("O7:R7")
            add_cell_value(
                worksheet2, 7, 15, 'Electrons delocalized by atom A to atom B', **tranparent_heading_properties)
            worksheet2.merge_cells("J6:R6")
            add_cell_value(
                worksheet2, 6, 10, 'Electron counts are printed in the descending order', **colored_heading_properties)
            worksheet2.merge_cells("U6:AE6")
            add_cell_value(
                worksheet2, 6, 21, 'The number of electrons delocalized by atom A to each atom of a molecular system, atom B', **tranparent_heading_properties)
            worksheet2.merge_cells(start_row=number_of_atoms+11,
                                   start_column=21, end_row=number_of_atoms+11, end_column=31)
            worksheet2.merge_cells(start_row=number_of_atoms+12,
                                   start_column=21, end_row=number_of_atoms+12, end_column=31)

            add_cell_value(worksheet2, number_of_atoms+11, 21,
                           "Values are printed in the descending order", **colored_heading_properties)
            add_cell_value(worksheet2, number_of_atoms+12, 21,
                           "The number of electrons delocalized by atom A to each atom of a molecular system, atom B", **tranparent_heading_properties)

            worksheet2.merge_cells(start_row=cell_row3+1, start_column=21,
                                   end_row=cell_row3+1, end_column=31)
            add_cell_value(worksheet2, cell_row3+1, 21,
                           "The number of electrons delocalized to atom A by each atom of a molecular system, atom B", **tranparent_heading_properties)

            worksheet2.merge_cells(start_row=cell_row4+1, start_column=21,
                                   end_row=cell_row4+1, end_column=31)
            add_cell_value(worksheet2, cell_row4+1, 21,
                           "The number of electrons delocalized to atom A by each atom of a molecular system, atom B", **tranparent_heading_properties)

            worksheet2.merge_cells(
                start_row=cell_row4, start_column=21, end_row=cell_row4, end_column=31)
            add_cell_value(worksheet2, cell_row4, 21,
                           "Values are printed in the descending order", **colored_heading_properties)

            # worksheet2.merge_cells(
            #     start_row=cell_row5+1, start_column=21, end_row=cell_row5+1, end_column=25)
            # add_cell_value(worksheet2, cell_row5+1, 21,
            #                "Electrons shared between atom A and entire molecule / system", **colored_heading_properties)

            # worksheet2.merge_cells(
            #     start_row=cell_row5+2, start_column=21, end_row=cell_row5+2, end_column=22)
            # add_cell_value(worksheet2, cell_row5+2, 21,
            #                "Atom-molecule pairs", **tranparent_heading_properties)
            # add_cell_value(worksheet2, cell_row5+2, 23,
            #                "e-shared", **tranparent_heading_properties)
            # worksheet2.merge_cells(
            #     start_row=cell_row5+2, start_column=24, end_row=cell_row5+2, end_column=25)
            # add_cell_value(worksheet2, cell_row5+2, 24,
            #                "e-delocalized by:", **tranparent_heading_properties)

            for range_string in ["C3:R4", "C7:E7", "G7:H7", "J7:M7", "F7:F7", "O7:R7", "U6:AE6", "U"+str(number_of_atoms+12)+":AE"+str(number_of_atoms+12), "U"+str(cell_row3+1)+":AE"+str(cell_row3+1), "U"+str(cell_row4+1)+":AE"+str(cell_row4+1)]:
                cells_range = worksheet2[range_string]
                for row in cells_range:
                    for cell in row:
                        cell.style = heading_style

            worksheet2["C3"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

            #                **heading_properties)
            # add_cell_value(worksheet2, 7, 5, 'e-shared', **heading_properties)
            # add_cell_value(worksheet2, 7, 12, 'e-shared', **heading_properties)
            # add_cell_value(worksheet2, 7, 16, 'e-delocalized by:',
            #                **heading_properties)
            # add_cell_value(
            #     worksheet2, 6, 10, 'Values of shared and delocalized electrons are in the descending order', **heading_properties)
            # add_cell_value(
            #     worksheet2, 6, 19, 'The number of electrons delocalized by atom A to atom B', **heading_properties)
            # add_cell_value(worksheet2, number_of_atoms+12, 19,
            #                'in the descending order', **heading_properties)

            for i, name in enumerate(molecule_names):
                cell = worksheet2.cell(row=2*number_of_atoms+17+i, column=21)
                cell.value = "Molecule-"+str(i+1)+" : "
                cell.font = Font(bold=True)
                cell = cell.offset(row=0, column=2)
                cell.value = name

            cell = worksheet2.cell(
                row=2*number_of_atoms+19+number_of_molecules, column=21)
            cell.value = "Molecular system:"
            cell.font = Font(bold=True)
            cell = cell.offset(row=0, column=3)
            cell.value = system_name

            molu_cnt = 1
            # molecule_names_new = molecule_names.copy()
            # del molecule_names_new[1]

            # for i, val in enumerate(new_list):
            #     cell = worksheet2.cell(row=9+i, column=20)
            #     if val != "":
            #         cell.value = "Molecule-"+str(molu_cnt)+" :"
            #         cell.font = Font(bold=True)
            #         molu_cnt += 1

            if image_filename == "" or image_filename == None:
                pass
            else:
                insert_image(image_filename, worksheet2,
                             'U'+str(2*number_of_atoms+21+number_of_molecules))

            # for merged_cell in worksheet2.merged_cells:
            #     merged_cell.shift(1, 0)

                # change the width of all the columns
            for col in worksheet2.columns:
                key = col[0].column_letter
                worksheet2.column_dimensions[key].width = 11.78





            doing_msg.set_progress(
                0.5, text_="Working on atoms-molecule-pairs sheet", speed=50)




            worksheet6 = workbook.create_sheet("atoms-molecule-pairs")


            headers = ["Pair count","Atom A", "Molecule-#",
                       "(A,Mol)", "A to Mol", "Mol to A"]


            for col_idx, header in enumerate(headers):
                cell = worksheet6.cell(row=8, column=col_idx+3)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                if col_idx > 1:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

             # write data rows for each atom
            row_idx = 9
            for row_id, values in enumerate(atom_molecule_dict.items(), start=9):
                if "_System" not in values[0]:
                    atom, molecule_ = values[0].split('_')
                    molecule = int(molecule_.split("-")[1])

                    cell = worksheet6.cell(row=row_idx, column=3)
                    cell.value = row_idx-8
                    cell = self.next_cell(cell)
                    cell.value = atom
                    cell = self.next_cell(cell)
                    cell.value = molecule
                    cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                    cell = self.next_cell(cell)
                    for col_idx, value in enumerate(values[1]):
                        cell.value = value
                        cell.number_format = '0.0000'
                        cell = self.next_cell(cell)
                    row_idx+=1

            sorted_atom_molecule_total_dict = {}
            sorted_atom_molecule_atom_to_molecule_dict = {}
            sorted_atom_molecule_molecule_to_atom_dict = {}
            for key, value in atom_molecule_dict.items():
                sorted_atom_molecule_total_dict[key] = value[0]
                sorted_atom_molecule_atom_to_molecule_dict[key] = value[1]
                sorted_atom_molecule_molecule_to_atom_dict[key] = value[2]

            sorted_atom_molecule_total_dict = dict(
                sorted(sorted_atom_molecule_total_dict.items(), key=lambda item: item[1], reverse=True))
            sorted_atom_molecule_atom_to_molecule_dict = dict(
                sorted(sorted_atom_molecule_atom_to_molecule_dict.items(), key=lambda item: item[1], reverse=True))
            sorted_atom_molecule_molecule_to_atom_dict = dict(
                sorted(sorted_atom_molecule_molecule_to_atom_dict.items(), key=lambda item: item[1], reverse=True))




            headers = ["Pair count", "Atom A", "Molecule-#",
                       "(A,Mol)", "", "Pair count", "Atom A", "Molecule-#", "A to Mol","","Pair count", "Atom A", "Molecule-#", "Mol to A"]


            for col_idx, header in enumerate(headers):
                if col_idx != 4 and col_idx != 9:
                    cell = worksheet6.cell(row=8, column=col_idx+10)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    if col_idx == 3 or col_idx == 8 or col_idx == 13:
                        cell.alignment = Alignment(
                            horizontal='center', vertical='center')


            # write data rows for each atom
            row_idx = 9
            for row_id, values in enumerate(sorted_atom_molecule_total_dict.items(), start=9):
                if "_System" not in values[0]:
                    atom, molecule_ = values[0].split('_')
                    molecule = int(molecule_.split("-")[1])


                    cell = worksheet6.cell(row=row_idx, column=10)
                    cell.value = row_idx-8
                    cell = self.next_cell(cell)
                    cell.value = atom
                    cell = self.next_cell(cell)
                    cell.value = molecule
                    cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                    cell = self.next_cell(cell)

                    cell.value = values[1]
                    cell.number_format = '0.0000'
                    cell = self.next_cell(cell)
                    row_idx+=1
            row_idx = 9

            for row_id, values in enumerate(sorted_atom_molecule_atom_to_molecule_dict.items(), start=9):
                if "_System" not in values[0]:
                    atom, molecule_ = values[0].split('_')
                    molecule = int(molecule_.split("-")[1])

                    cell = worksheet6.cell(row=row_idx, column=15)
                    cell.value = row_idx-8
                    cell = self.next_cell(cell)
                    cell.value = atom
                    cell = self.next_cell(cell)
                    cell.value = molecule
                    cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                    cell = self.next_cell(cell)

                    cell.value = values[1]
                    cell.number_format = '0.0000'
                    cell = self.next_cell(cell)
                    row_idx+=1

            row_idx = 9
            for row_id, values in enumerate(sorted_atom_molecule_molecule_to_atom_dict.items(), start=9):
                if "_System" not in values[0]:
                    atom, molecule_ = values[0].split('_')
                    molecule = int(molecule_.split("-")[1])

                    cell = worksheet6.cell(row=row_idx, column=20)
                    cell.value = row_idx-8
                    cell = self.next_cell(cell)
                    cell.value = atom
                    cell = self.next_cell(cell)
                    cell.value = molecule
                    cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                    cell = self.next_cell(cell)

                    cell.value = values[1]
                    cell.number_format = '0.0000'
                    cell = self.next_cell(cell)
                    row_idx+=1

            new_row_idx = len(atom_molecule_dict)+12+28-number_of_atoms
            if number_of_molecules > 1:


                headers = ["Pair count","Atom A","(A,MS)", "A to MS", "MS to A"]


                for col_idx, header in enumerate(headers):
                    cell = worksheet6.cell(row=new_row_idx-1, column=col_idx+3)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    if col_idx > 1:
                        cell.alignment = Alignment(
                            horizontal='center', vertical='center')

                # write data rows for each atom
                row_idx = new_row_idx
                for row_id, values in enumerate(atom_molecule_dict.items(), start=9):
                    if "_System" in values[0]:
                        atom, molecule_ = values[0].split('_')

                        cell = worksheet6.cell(row=row_idx, column=3)
                        cell.value = row_idx-new_row_idx+1
                        cell = self.next_cell(cell)
                        cell.value = atom
                        cell = self.next_cell(cell)
                        for col_idx, value in enumerate(values[1]):
                            cell.value = value
                            cell.number_format = '0.0000'
                            cell = self.next_cell(cell)
                        row_idx+=1


                headers = ["Pair count", "Atom A","(A,MS)", "" ,"", "Pair count", "Atom A", "A to MS","","","Pair count", "Atom A", "MS to A"]


                for col_idx, header in enumerate(headers):
                    if header != "":
                        cell = worksheet6.cell(row=new_row_idx-1, column=col_idx+10)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                        if col_idx == 2 or col_idx == 7 or col_idx == 12:
                            cell.alignment = Alignment(
                                horizontal='center', vertical='center')


                # write data rows for each atom
                row_idx = new_row_idx
                for row_id, values in enumerate(sorted_atom_molecule_total_dict.items(), start=9):
                    if "_System" in values[0]:
                        atom, molecule_ = values[0].split('_')


                        cell = worksheet6.cell(row=row_idx, column=10)
                        cell.value = row_idx-new_row_idx+1
                        cell = self.next_cell(cell)
                        cell.value = atom
                        cell = self.next_cell(cell)

                        cell.value = values[1]
                        cell.number_format = '0.0000'
                        cell = self.next_cell(cell)
                        row_idx+=1

                row_idx = new_row_idx
                for row_id, values in enumerate(sorted_atom_molecule_atom_to_molecule_dict.items(), start=9):
                    if "_System" in values[0]:
                        atom, molecule_ = values[0].split('_')

                        cell = worksheet6.cell(row=row_idx, column=15)
                        cell.value = row_idx-new_row_idx+1
                        cell = self.next_cell(cell)
                        cell.value = atom
                        cell = self.next_cell(cell)
                        cell.value = values[1]
                        cell.number_format = '0.0000'
                        cell = self.next_cell(cell)
                        row_idx+=1

                row_idx = new_row_idx
                for row_id, values in enumerate(sorted_atom_molecule_molecule_to_atom_dict.items(), start=9):
                    if "_System" in values[0]:
                        atom, molecule_ = values[0].split('_')


                        cell = worksheet6.cell(row=row_idx, column=20)
                        cell.value = row_idx-new_row_idx+1
                        cell = self.next_cell(cell)
                        cell.value = atom
                        cell = self.next_cell(cell)

                        cell.value = values[1]
                        cell.number_format = '0.0000'
                        cell = self.next_cell(cell)
                        row_idx+=1



                worksheet6.merge_cells('C{0}:D{0}'.format(new_row_idx-2))
                add_cell_value(worksheet6, new_row_idx-2, 3, 'All atoms-MS pairs', **tranparent_heading_properties)

                add_cell_value(worksheet6, new_row_idx-2, 5, 'e-shared', **tranparent_heading_properties)

                worksheet6.merge_cells('F{0}:G{0}'.format(new_row_idx-2))
                add_cell_value(worksheet6, new_row_idx-2, 6, 'e-delocalized by:', **tranparent_heading_properties)

                worksheet6.merge_cells('J{0}:L{0}'.format(new_row_idx-2))
                add_cell_value(worksheet6, new_row_idx-2, 10, 'Electrons shared by atom A and MS', **tranparent_heading_properties)

                worksheet6.merge_cells('O{0}:Q{0}'.format(new_row_idx-2))
                add_cell_value(worksheet6, new_row_idx-2, 15, 'Electrons delocalized by atom A to MS', **tranparent_heading_properties)

                worksheet6.merge_cells('T{0}:V{0}'.format(new_row_idx-2))
                add_cell_value(worksheet6, new_row_idx-2, 20, 'Electrons delocalized by MS to atom A', **tranparent_heading_properties)

                worksheet6.merge_cells('J{0}:V{0}'.format(new_row_idx-3))
                add_cell_value(worksheet6, new_row_idx-3, 10, 'Electron counts are printed in the descending order', **colored_heading_properties)


                for range_string in ['C{0}:D{0}'.format(new_row_idx-2),"E{0}:E{0}".format(new_row_idx-2),'F{0}:G{0}'.format(new_row_idx-2),'J{0}:L{0}'.format(new_row_idx-2),'O{0}:Q{0}'.format(new_row_idx-2),'T{0}:V{0}'.format(new_row_idx-2)]:
                    cells_range = worksheet6[range_string]
                    for row in cells_range:
                        for cell in row:
                            cell.style = heading_style


            worksheet6.merge_cells('C3:W4')
            add_cell_value(worksheet6, 3, 3, '''The number of electrons that are (i) shared by atom-molecule or atom-(molecular system, MS) pairs is shown as 'e-shared' by (A,Mol) or (A,MS), whereas electrons delocalozed (ii) by atom A to either molecule or MS and (iii) delocalized by either a  molecule or MS to atom A are shown as 'e-delocalized by:'. These values are computed for all unique atom-molecule (A,Mol) and atom-MS (A,MS) pairs of a molecular system.''', **tranparent_heading_properties)

            worksheet6.merge_cells('C7:E7')
            add_cell_value(worksheet6, 7, 3, 'All atoms-molecule pairs', **tranparent_heading_properties)

            add_cell_value(worksheet6, 7, 6, 'e-shared', **tranparent_heading_properties)


            worksheet6.merge_cells('G7:H7')
            add_cell_value(worksheet6, 7, 7, 'e-delocalized by:', **tranparent_heading_properties)

            worksheet6.merge_cells('J7:M7')
            add_cell_value(worksheet6, 7, 10, 'Electrons shared by atom A and molecule', **tranparent_heading_properties)


            worksheet6.merge_cells('O7:R7')
            add_cell_value(worksheet6, 7, 15, 'Electrons delocalized by atom A to molecule', **tranparent_heading_properties)

            worksheet6.merge_cells('T7:W7')
            add_cell_value(worksheet6, 7, 20, 'Electrons delocalized by molecule to atom A', **tranparent_heading_properties)



            worksheet6.merge_cells('J6:W6')
            add_cell_value(worksheet6, 6, 10, 'Electron counts are printed in the descending order', **colored_heading_properties)



            for range_string in ["C3:W4",'C7:E7', "F7:F7","G7:H7", "J7:M7", "O7:R7", "T7:W7"]:
                cells_range = worksheet6[range_string]
                for row in cells_range:
                    for cell in row:
                        cell.style = heading_style

            worksheet6["C3"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)









            if image_filename == "" or image_filename == None:
                print("No Image selected")
            else:
                insert_image(image_filename, worksheet6,
                             'C'+str(len(atom_molecule_dict)+14-number_of_atoms))

            cell = worksheet6.cell(row=len(atom_molecule_dict)+12-number_of_atoms, column=3)
            cell.value = "Molecular system:"
            cell.font = Font(bold=True)
            cell = cell.offset(row=0, column=3)
            cell.value = system_name

            for i, name in enumerate(molecule_names):
                cell = worksheet6.cell(row=len(atom_molecule_dict)+12+i-number_of_atoms, column=10)
                cell.value = "Molecule-"+str(i+1)+" : "
                cell.font = Font(bold=True)
                cell = cell.offset(row=0, column=2)
                cell.value = name




            for col in worksheet6.columns:
                key = col[0].column_letter
                worksheet6.column_dimensions[key].width = 11.78














            doing_msg.set_progress(
                0.6, text_="Working on intra-Mol(A,B)_cov sheet", speed=50)



            worksheet3 = workbook.create_sheet("intra-Mol(A,B)_cov")

            # create header cell for intramolecular electron delocalization information in worksheet3 table

            mole_name_properties = {
                'font': Font(bold=True, color='FF0000'),
            }
            headers = ["(A,B) count", "Atom A",
                       "Atom B", "(A,B)", "A to B", "B to A"]
            headers_properties = {
                'font': header_font,
                'alignment': openpyxl.styles.Alignment(horizontal='center', vertical='center'),
                "fill": header_fill
            }
            headers_properties_LEFT = {
                'font': header_font,

                "fill": header_fill
            }
            data_properties = {
                "number_format": "0.0000",
            }
            percentage_properties = {
                "number_format": "0.00",
            }

            row_idx = 7
            for name in molecule_names:
                mol_name = name
                cov_list = intra_mol_dict[name]["covalent"]
                add_cell_value(worksheet3, row_idx-1, 2,
                               mol_name, **mole_name_properties)
                worksheet3.merge_cells(
                    start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
                add_cell_value(
                    worksheet3, row_idx, 2, "Covalently bonded atom-pairs (A,B)", **mole_name_properties)
                worksheet3.merge_cells(
                    start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
                add_cell_value(worksheet3, row_idx, 5,
                               "e-shared", **heading_properties)
                add_cell_value(worksheet3, row_idx, 6,
                               "e-delocalized by:", **heading_properties)
                # worksheet3.merge_cells(
                #     start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)

                for range_string in ["B"+str(row_idx)+":D"+str(row_idx), "F"+str(row_idx)+":G"+str(row_idx), "E"+str(row_idx)+":E"+str(row_idx)]:
                    cells_range = worksheet3[range_string]
                    for row in cells_range:
                        for cell in row:
                            cell.style = heading_style

                row_idx += 1
                for idx, text in enumerate(headers):
                    if idx <= 2:
                        add_cell_value(worksheet3, row_idx, idx+2,
                                       text, **headers_properties_LEFT)
                    else:
                        add_cell_value(worksheet3, row_idx, idx+2,
                                       text, **headers_properties)
                row_idx += 1
                for idx, values in enumerate(cov_list.items(), start=1):
                    add_cell_value(worksheet3, row_idx, 2, idx)

                    A, B = self.get_atom_pair(values[0])
                    add_cell_value(worksheet3, row_idx, 3, A)
                    add_cell_value(worksheet3, row_idx, 4, B)
                    add_cell_value(worksheet3, row_idx, 5,
                                   values[1][0], **data_properties)
                    add_cell_value(worksheet3, row_idx, 6,
                                   values[1][1], **data_properties)
                    add_cell_value(worksheet3, row_idx, 7,
                                   values[1][2], **data_properties)
                    row_idx += 1

                add_cell_value(worksheet3, row_idx, 4,
                               "Total:", **total_properties)
                add_cell_value(worksheet3, row_idx, 5, sum(
                    [i[1][0] for i in cov_list.items()]), **number_properties)
                add_cell_value(worksheet3, row_idx, 6, sum(
                    [i[1][1] for i in cov_list.items()]), **number_properties)
                add_cell_value(worksheet3, row_idx, 7, sum(
                    [i[1][2] for i in cov_list.items()]), **number_properties)
                row_idx += 3

            new_row_idx = row_idx+5

            # write the header row
            headers = ["(A,B) count", "Atom A", "Atom B",
                       "(A,B)", "", "(A,B) count", "Atom A", "Atom B", "A to B"]
            for col_idx, header in enumerate(headers):
                if col_idx != 4:
                    if col_idx == 3 or col_idx == 8:
                        add_cell_value(worksheet3, 8, col_idx+9,
                                       header, **headers_properties)
                        print(header)
                    else:
                        add_cell_value(worksheet3, 8, col_idx+9,
                                       header, **headers_properties_LEFT)

            for row_idx, values in enumerate(intra_sorted_according_to_total_deloc_electron["covalent"].items(), start=9):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-8), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=9):
                    cell = worksheet3.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 9:
                        cell.number_format = '0.0000'

            for row_idx, values in enumerate(intra_sorted_according_to_contri["covalent"].items(), start=9):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-8), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=14):
                    cell = worksheet3.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 14:
                        cell.number_format = '0.0000'

            add_cell_value(worksheet3, len(
                intra_sorted_according_to_total_deloc_electron["covalent"])+9, 11, "Total:", **total_properties)
            add_cell_value(worksheet3, len(intra_sorted_according_to_total_deloc_electron["covalent"])+9, 12, sum(
                [i[1] for i in intra_sorted_according_to_total_deloc_electron["covalent"].items()]), **number_properties)

            add_cell_value(worksheet3, len(
                intra_sorted_according_to_contri["covalent"])+9, 16, "Total:", **total_properties)
            add_cell_value(worksheet3, len(intra_sorted_according_to_contri["covalent"])+9, 17, sum(
                [i[1] for i in intra_sorted_according_to_contri["covalent"].items()]), **number_properties)

            col_idx = 20
            temp = 0
            for key, datas in data_of_table2_intra["covalent"].items():

                add_cell_value(worksheet3, 8, col_idx,
                               "To atom B", **headers_properties_LEFT)
                add_cell_value(worksheet3, 7, col_idx+1,
                               "By atom A", **headers_properties)
                add_cell_value(worksheet3, 8, col_idx+1,
                               key, **headers_properties)

                for row_idx, value in enumerate(datas.items(), start=9):

                    add_cell_value(worksheet3, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx +
                                   1, value[1], **data_properties)

                add_cell_value(worksheet3, len(datas)+9, col_idx,
                               "Total:", **total_properties)
                temp = max(temp, len(datas)+9)
                add_cell_value(worksheet3, len(datas)+9, col_idx+1,
                               sum([i[1] for i in datas.items()]), **number_properties)

                col_idx += 4

            col_idx = 20
            row_add = temp-3
            cell_row2 = 7 + row_add
            for key, datas in sorted_data_of_table2_intra["covalent"].items():

                add_cell_value(worksheet3, 9 + row_add, col_idx,
                               "To atom B", **headers_properties_LEFT)
                add_cell_value(worksheet3, 8 + row_add, col_idx +
                               1, "By atom A", **headers_properties)
                add_cell_value(worksheet3, 9 + row_add, col_idx +
                               1, key, **headers_properties)
                add_cell_value(worksheet3, 9 + row_add, col_idx +
                               2, '''%-fraction''', **headers_properties)


                tot = sum([i[1] for i in datas.items()])
                for row_idx, value in enumerate(datas.items(), start=10 + row_add):

                    add_cell_value(worksheet3, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx +
                                   1, value[1], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx +
                                      2, value[1]*100/tot, **percentage_properties)


                add_cell_value(worksheet3, len(datas)+10 + row_add,
                               col_idx, "Total:", **total_properties)
                temp = max(temp, len(datas)+10 + row_add)
                add_cell_value(worksheet3, len(datas)+10 + row_add, col_idx+1,
                               sum([i[1] for i in datas.items()]), **number_properties)
                add_cell_value(worksheet3, len(datas)+10 + row_add, col_idx+2,
                                100, **total_properties)

                col_idx += 4


            for  key, datas in data_of_table2_intra["covalent"].items():
                for key2, value in datas.items():

                    reversed_data_of_table2_intra_cov[key2][key] = value

            reversed_sorted_data_of_table2_intra_cov = {}
            for key, datas in reversed_data_of_table2_intra_cov.items():
                reversed_sorted_data_of_table2_intra_cov[key] = dict(sorted(datas.items(), key=lambda item: item[1], reverse=True))

            col_idx = 20
            row_add = temp+23+number_of_molecules
            cell_row3 = 7 + row_add
            temp_ = 0
            for key, datas in reversed_data_of_table2_intra_cov.items():

                add_cell_value(worksheet3, 9+row_add, col_idx,
                                 "By atom B", **headers_properties_LEFT)
                add_cell_value(worksheet3, 8+row_add, col_idx+1,
                                    "To atom A", **headers_properties)
                add_cell_value(worksheet3, 9+row_add, col_idx+1,
                                    key, **headers_properties)

                for row_idx, value in enumerate(datas.items(), start=10+row_add):

                    add_cell_value(worksheet3, row_idx, col_idx,
                                    value[0], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx+1,
                                    value[1], **data_properties)

                add_cell_value(worksheet3, len(datas)+10+row_add, col_idx,
                                    "Total:", **total_properties)
                temp_ = max(temp_, len(datas)+10+row_add)
                add_cell_value(worksheet3, len(datas)+10+row_add, col_idx+1,
                                    sum([i[1] for i in datas.items()]), **number_properties)

                col_idx += 4

            col_idx = 20
            row_add = temp_-3
            cell_row4 = 7 + row_add
            temp_ = 0

            for key, datas in reversed_sorted_data_of_table2_intra_cov.items():

                add_cell_value(worksheet3, 9+row_add, col_idx,
                                "By atom B", **headers_properties_LEFT)
                add_cell_value(worksheet3, 8+row_add, col_idx+1,
                                    "To atom A", **headers_properties)
                add_cell_value(worksheet3, 9+row_add, col_idx+1,
                                    key, **headers_properties)

                add_cell_value(worksheet3, 9+row_add, col_idx+2,
                                 '''%-fraction''', **headers_properties)



                tot= sum([i[1] for i in datas.items()])
                for row_idx, value in enumerate(datas.items(), start=10+row_add):

                    add_cell_value(worksheet3, row_idx, col_idx,
                                    value[0], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx+1,
                                    value[1], **data_properties)
                    add_cell_value(worksheet3, row_idx, col_idx+2,
                                    value[1]*100/tot, **data_properties)

                add_cell_value(worksheet3, len(datas)+10+row_add, col_idx,
                                    "Total:", **total_properties)
                temp_ = max(temp_, len(datas)+10+row_add)
                add_cell_value(worksheet3, len(datas)+10+row_add, col_idx+1,
                                    sum([i[1] for i in datas.items()]), **number_properties)
                add_cell_value(worksheet3, len(datas)+10+row_add, col_idx+2,
                                    100, **total_properties)

                col_idx += 4



            worksheet3.insert_cols(1, 1)
            for merged_cell in worksheet3.merged_cells:
                merged_cell.shift(1, 0)

            worksheet3.merge_cells("C3:R4")
            add_cell_value(worksheet3, 3, 3, '''The number of electrons (i) shared by atom-piar (A,B) is shown as 'e-shared', (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A are shown as 'e-delocalized by:'.                                                                                            These values are computed for all unique atom-pairs (A,B) of a molecular system that are considered as being covalently bonded according to a Lewis dogma.''', **tranparent_heading_properties)
            # worksheet3.merge_cells("C7:E7")
            # add_cell_value(worksheet3, 7, 3, 'All unique atom-pairs (A,B)', **tranparent_heading_properties)
            # worksheet3.merge_cells("G7:H7")
            # add_cell_value(worksheet3, 7, 6, 'e-shared', **tranparent_heading_properties)
            # add_cell_value(worksheet3, 7, 7, 'e-delocalized by:', **tranparent_heading_properties)
            worksheet3.merge_cells("J7:M7")
            add_cell_value(
                worksheet3, 7, 10, 'Electons shared by atoms A and B', **tranparent_heading_properties)
            worksheet3.merge_cells("O7:R7")
            add_cell_value(
                worksheet3, 7, 15, 'Electrons delocalized by atom A to atom B', **tranparent_heading_properties)
            worksheet3.merge_cells("J6:R6")
            add_cell_value(
                worksheet3, 6, 10, 'Shared and delocalized electron counts are printed in the descending order', **colored_heading_properties)
            worksheet3.merge_cells("U6:AE6")
            add_cell_value(
                worksheet3, 6, 21, 'The number of electrons delocalized by atom A to atom B that is considered as being covalently bonded to atom A', **tranparent_heading_properties)
            worksheet3.merge_cells(start_row=cell_row2,
                                   start_column=21, end_row=cell_row2, end_column=31)
            worksheet3.merge_cells(start_row=cell_row2-1,
                                   start_column=21, end_row=cell_row2-1, end_column=31)

            add_cell_value(worksheet3,  cell_row2-1, 21,
                           "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)
            add_cell_value(worksheet3, cell_row2 , 21,
                           "The number of electrons delocalized by atom A to atom B that is considered as being covalently bonded to atom A", **tranparent_heading_properties)

            worksheet3.merge_cells("F2:N2")
            add_cell_value(
                worksheet3, 2, 6, "The minimum threshold number of shared electrons by atoms classically considered as being covalently bonded is:", **tranparent_heading_properties)
            add_cell_value(worksheet3, 2, 15, cov_threshold,
                           **heading_properties)

            worksheet3.merge_cells(start_row=cell_row3,start_column=21, end_row=cell_row3, end_column=31)
            add_cell_value(worksheet3, cell_row3, 21, "The number of electrons delocalized to atom A by atom B that is considered as being covalently bonded to atom A", **tranparent_heading_properties)
            worksheet3.merge_cells(start_row=cell_row4,start_column=21, end_row=cell_row4, end_column=31)
            add_cell_value(worksheet3, cell_row4, 21, "The number of electrons delocalized to atom A by atom B that is considered as being covalently bonded to atom A", **tranparent_heading_properties)

            worksheet3.merge_cells(start_row=cell_row4-1,start_column=21, end_row=cell_row4-1, end_column=31)
            add_cell_value(worksheet3, cell_row4-1, 21, "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)




            for range_string in ["C3:R4", "J7:M7", "O7:R7", "U6:AE6", "U"+str(cell_row2)+":AE"+str(cell_row2), "U"+str(cell_row3)+":AE"+str(cell_row3), "U"+str(cell_row4)+":AE"+str(cell_row4)]:
                cells_range = worksheet3[range_string]
                for row in cells_range:
                    for cell in row:
                        cell.style = heading_style
            worksheet3["C3"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

            #                **heading_properties)
            # add_cell_value(worksheet3, 7, 5, 'e-shared', **heading_properties)
            # add_cell_value(worksheet3, 7, 12, 'e-shared', **heading_properties)
            # add_cell_value(worksheet3, 7, 16, 'e-delocalized by:',
            #                **heading_properties)
            # add_cell_value(
            #     worksheet3, 6, 10, 'Values of shared and delocalized electrons are in the descending order', **heading_properties)
            # add_cell_value(
            #     worksheet3, 6, 19, 'The number of electrons delocalized by atom A to atom B', **heading_properties)
            # add_cell_value(worksheet3, number_of_atoms+12, 19,
            #                'in the descending order', **heading_properties)

            for i, name in enumerate(molecule_names):
                cell = worksheet3.cell(row=temp+3+i, column=21)
                cell.value = "Molecule-"+str(i+1)+" : "
                cell.font = Font(bold=True)
                cell = cell.offset(row=0, column=2)
                cell.value = name

            cell = worksheet3.cell(row=temp+5+number_of_molecules, column=21)
            cell.value = "Molecular system:"
            cell.font = Font(bold=True)
            cell = cell.offset(row=0, column=3)
            cell.value = system_name

            # add_cell_value(worksheet3, 8, 12, "e-shared", **heading_properties)
            # worksheet3.merge_cells("P8:Q8")
            # add_cell_value(worksheet3, 8, 16, "e-delocalized by:",
            #                **heading_properties)

            # worksheet3.merge_cells("B1:G1")

            # add_cell_value(
            #     worksheet3, 1, 2, "The threshold number of shared electrons by (A,B) is set to:", **mole_name_properties)
            # add_cell_value(worksheet3, 1, 8, cov_threshold,
            #                **heading_properties)

            # worksheet3.merge_cells("B3:K4")
            # add_cell_value(worksheet3, 3, 2, "The number of electrons (i) shared by atom-piar (A,B) (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A computed for all unique intramolecular atom-pairs (A,B) of a molecular system", **heading_properties)

            # worksheet3.merge_cells("B7:D7")
            # add_cell_value(worksheet3, 7, 2,
            #                "(A,B) covalently bonded", **heading_properties)

            # worksheet3.merge_cells("J7:P7")
            # add_cell_value(
            #     worksheet3, 7, 10, "Values of shared and delocalized electrons are in the descending order", **heading_properties)

            # worksheet3.merge_cells("U7:AB7")
            # add_cell_value(
            #     worksheet3, 7, 21, "The number of electrons delocalized by atom A to covalently-bonded atom B", **heading_properties)

            # worksheet3.merge_cells(
            #     start_row=6 + row_add, start_column=21, end_row=6 + row_add, end_column=28)
            # add_cell_value(worksheet3, 6 + row_add, 21,
            #                "The number of electrons delocalized by atom A to covalently-bonded atom B", **heading_properties)

            # worksheet3.merge_cells(
            #     start_row=7+row_add, start_column=21, end_row=7+row_add, end_column=23)
            # add_cell_value(worksheet3, 7+row_add, 21,
            #                "in the descending order", **heading_properties)

            if image_filename == "" or image_filename == None:
                pass
            else:
                insert_image(image_filename, worksheet3,
                             'U'+str(temp+7+number_of_molecules))

            # change the width of all the columns
            for col in worksheet3.columns:
                key = col[0].column_letter
                worksheet3.column_dimensions[key].width = 11.78

            # Create a exactly same worksheet4 for non-covalent bond

            doing_msg.set_progress(
                0.75, "Working on intra-Mol(A,B)_non-cov sheet...", speed=40)

            worksheet4 = workbook.create_sheet("intra-Mol(A,B)_non-cov")

            # create header cell for intramolecular electron delocalization information in worksheet3 table

            mole_name_properties = {
                'font': Font(bold=True, color='FF0000'),
            }
            headers = ["(A,B) count", "Atom A",
                       "Atom B", "(A,B)", "A to B", "B to A"]
            headers_properties = {
                'font': header_font,
                'alignment': openpyxl.styles.Alignment(horizontal='center', vertical='center'),
                "fill": header_fill
            }
            headers_properties_LEFT = {
                'font': header_font,

                "fill": header_fill
            }
            data_properties = {
                "number_format": "0.0000",
            }


            row_idx = 7
            for name in molecule_names:
                mol_name = name
                cov_list = intra_mol_dict[name]["non-covalent"]
                add_cell_value(worksheet4, row_idx-1, 2,
                               mol_name, **mole_name_properties)
                worksheet4.merge_cells(
                    start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
                add_cell_value(
                    worksheet4, row_idx, 2, "Non-bonded atom-pairs (A,B)", **mole_name_properties)
                worksheet4.merge_cells(
                    start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
                add_cell_value(worksheet4, row_idx, 5,
                               "e-shared", **heading_properties)
                add_cell_value(worksheet4, row_idx, 6,
                               "e-delocalized by:", **heading_properties)
                # worksheet4.merge_cells(
                #     start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)

                for range_string in ["B"+str(row_idx)+":D"+str(row_idx), "F"+str(row_idx)+":G"+str(row_idx), "E"+str(row_idx)+":E"+str(row_idx)]:
                    cells_range = worksheet4[range_string]
                    for row in cells_range:
                        for cell in row:
                            cell.style = heading_style

                row_idx += 1
                for idx, text in enumerate(headers):
                    if idx <= 2:
                        add_cell_value(worksheet4, row_idx, idx+2,
                                       text, **headers_properties_LEFT)
                    else:
                        add_cell_value(worksheet4, row_idx, idx+2,
                                       text, **headers_properties)
                row_idx += 1
                for idx, values in enumerate(cov_list.items(), start=1):
                    add_cell_value(worksheet4, row_idx, 2, idx)

                    A, B = self.get_atom_pair(values[0])
                    add_cell_value(worksheet4, row_idx, 3, A)
                    add_cell_value(worksheet4, row_idx, 4, B)
                    add_cell_value(worksheet4, row_idx, 5,
                                   values[1][0], **data_properties)
                    add_cell_value(worksheet4, row_idx, 6,
                                   values[1][1], **data_properties)
                    add_cell_value(worksheet4, row_idx, 7,
                                   values[1][2], **data_properties)
                    row_idx += 1

                add_cell_value(worksheet4, row_idx, 4,
                               "Total:", **total_properties)
                add_cell_value(worksheet4, row_idx, 5, sum(
                    [i[1][0] for i in cov_list.items()]), **number_properties)
                add_cell_value(worksheet4, row_idx, 6, sum(
                    [i[1][1] for i in cov_list.items()]), **number_properties)
                add_cell_value(worksheet4, row_idx, 7, sum(
                    [i[1][2] for i in cov_list.items()]), **number_properties)
                row_idx += 3

            new_row_idx = row_idx+5

            # write the header row
            headers = ["(A,B) count", "Atom A", "Atom B",
                       "(A,B)", "", "(A,B) count", "Atom A", "Atom B", "A to B"]
            for col_idx, header in enumerate(headers):
                if col_idx != 4:
                    if col_idx == 3 or col_idx == 8:
                        add_cell_value(worksheet4, 8, col_idx+9,
                                       header, **headers_properties)
                        print(header)
                    else:
                        add_cell_value(worksheet4, 8, col_idx+9,
                                       header, **headers_properties_LEFT)

            for row_idx, values in enumerate(intra_sorted_according_to_total_deloc_electron["non-covalent"].items(), start=9):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-8), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=9):
                    cell = worksheet4.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 9:
                        cell.number_format = '0.0000'

            for row_idx, values in enumerate(intra_sorted_according_to_contri["non-covalent"].items(), start=9):
                _1, _2 = self.get_atom_pair(values[0])
                new_values = [int(row_idx-8), _1, _2, values[1]]

                for col_idx, value in enumerate(new_values, start=14):
                    cell = worksheet4.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    if col_idx != 14:
                        cell.number_format = '0.0000'

            add_cell_value(worksheet4, len(
                intra_sorted_according_to_total_deloc_electron["non-covalent"])+9, 11, "Total:", **total_properties)
            add_cell_value(worksheet4, len(intra_sorted_according_to_total_deloc_electron["non-covalent"])+9, 12, sum(
                [i[1] for i in intra_sorted_according_to_total_deloc_electron["non-covalent"].items()]), **number_properties)

            add_cell_value(worksheet4, len(
                intra_sorted_according_to_contri["non-covalent"])+9, 16, "Total:", **total_properties)
            add_cell_value(worksheet4, len(intra_sorted_according_to_contri["non-covalent"])+9, 17, sum(
                [i[1] for i in intra_sorted_according_to_contri["non-covalent"].items()]), **number_properties)

            col_idx = 20
            temp = 0
            for key, datas in data_of_table2_intra["non-covalent"].items():

                add_cell_value(worksheet4, 8, col_idx,
                               "To atom B", **headers_properties_LEFT)
                add_cell_value(worksheet4, 7, col_idx+1,
                               "By atom A", **headers_properties)
                add_cell_value(worksheet4, 8, col_idx+1,
                               key, **headers_properties)

                for row_idx, value in enumerate(datas.items(), start=9):

                    add_cell_value(worksheet4, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx +
                                   1, value[1], **data_properties)

                add_cell_value(worksheet4, len(datas)+9, col_idx,
                               "Total:", **total_properties)
                temp = max(temp, len(datas)+9)
                add_cell_value(worksheet4, len(datas)+9, col_idx+1,
                               sum([i[1] for i in datas.items()]), **number_properties)

                col_idx += 4

            col_idx = 20
            row_add = temp-3
            cell_row2 = row_add
            for key, datas in sorted_data_of_table2_intra["non-covalent"].items():

                add_cell_value(worksheet4, 9 + row_add, col_idx,
                               "To atom B", **headers_properties_LEFT)
                add_cell_value(worksheet4, 8 + row_add, col_idx +
                               1, "By atom A", **headers_properties)
                add_cell_value(worksheet4, 9 + row_add, col_idx +
                               1, key, **headers_properties)
                add_cell_value(worksheet4, 9 + row_add, col_idx +
                               2, '''%-fraction''', **headers_properties)

                tot = sum([i[1] for i in datas.items()])
                for row_idx, value in enumerate(datas.items(), start=10 + row_add):

                    add_cell_value(worksheet4, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx +
                                   1, value[1], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx +
                                   2, value[1]*100/tot, **percentage_properties)

                add_cell_value(worksheet4, len(datas)+10 + row_add,
                               col_idx, "Total:", **total_properties)
                temp = max(temp, len(datas)+10 + row_add)
                add_cell_value(worksheet4, len(datas)+10 + row_add, col_idx+1,
                               sum([i[1] for i in datas.items()]), **number_properties)
                add_cell_value(worksheet4, len(datas)+10 + row_add,
                               col_idx+2, 100, **total_properties)
                col_idx += 4


            for key, datas in data_of_table2_intra["non-covalent"].items():
                for key2, value in datas.items():
                    if key2 not in reversed_data_of_table2_intra:
                        reversed_data_of_table2_intra[key2] = {}
                    reversed_data_of_table2_intra[key2][key] = value

            reversed_sorted_data_of_table2_intra = {}
            for key, datas in reversed_data_of_table2_intra.items():
                reversed_sorted_data_of_table2_intra[key] = dict(
                    sorted(datas.items(), key=lambda item: item[1], reverse=True))

            col_idx = 20
            row_add = temp+25+number_of_molecules
            cell_row3 = row_add+7
            for key, datas in reversed_data_of_table2_intra.items():
                add_cell_value(worksheet4, 9 + row_add, col_idx,
                               "By atom B", **headers_properties_LEFT)
                add_cell_value(worksheet4, 8 + row_add, col_idx +
                               1, "To atom A", **headers_properties)
                add_cell_value(worksheet4, 9 + row_add,
                               col_idx + 1, key, **headers_properties)

                for row_idx, value in enumerate(datas.items(), start=10 + row_add):
                    add_cell_value(worksheet4, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx +
                                   1, value[1], **data_properties)

                add_cell_value(worksheet4, len(datas) + 10 +
                               row_add, col_idx, "Total:", **total_properties)
                temp_temp = max(temp, len(datas) + 10 + row_add)
                add_cell_value(worksheet4, len(datas) + 10 + row_add, col_idx + 1,
                               sum([i[1] for i in datas.items()]), **number_properties)

                col_idx += 4

            col_idx = 20
            row_add = temp_temp-2
            cell_row4 = row_add+7

            for key, datas in reversed_sorted_data_of_table2_intra.items():

                add_cell_value(worksheet4, 9 + row_add, col_idx,
                               "By atom B", **headers_properties_LEFT)
                add_cell_value(worksheet4, 8 + row_add, col_idx +
                               1, "To atom A", **headers_properties)
                add_cell_value(worksheet4, 9 + row_add,
                               col_idx + 1, key, **headers_properties)

                add_cell_value(worksheet4, 9 + row_add, col_idx + 2,
                               '''%-fraction''', **headers_properties)

                tot = sum([i[1] for i in datas.items()])
                for row_idx, value in enumerate(datas.items(), start=10 + row_add):
                    add_cell_value(worksheet4, row_idx, col_idx,
                                   value[0], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx +
                                   1, value[1], **data_properties)
                    add_cell_value(worksheet4, row_idx, col_idx + 2,
                                   value[1] * 100 / tot, **percentage_properties)

                add_cell_value(worksheet4, len(datas) + 10 +
                               row_add, col_idx, "Total:", **total_properties)

                add_cell_value(worksheet4, len(datas) + 10 + row_add, col_idx + 1,
                               sum([i[1] for i in datas.items()]), **number_properties)
                add_cell_value(worksheet4, len(datas) + 10 +
                               row_add, col_idx + 2, 100, **total_properties)

                col_idx += 4

            worksheet4.insert_cols(1, 1)
            for merged_cell in worksheet4.merged_cells:
                merged_cell.shift(1, 0)

            worksheet4.merge_cells("C3:R4")
            add_cell_value(worksheet4, 3, 3, '''The number of electrons (i) shared by atom-piar (A,B) is shown as 'e-shared', (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A are shown as 'e-delocalized by:'.                                                                                                  These values are computed for all unique atom-pairs (A,B) of a molecular system that are considered as being involved in non-covalent interactions, hence no covalent bond links them.''', **tranparent_heading_properties)
            # worksheet4.merge_cells("C7:E7")
            # add_cell_value(worksheet4, 7, 3, 'All unique atom-pairs (A,B)', **tranparent_heading_properties)
            # worksheet4.merge_cells("G7:H7")
            # add_cell_value(worksheet4, 7, 6, 'e-shared', **tranparent_heading_properties)
            # add_cell_value(worksheet4, 7, 7, 'e-delocalized by:', **tranparent_heading_properties)
            worksheet4.merge_cells("J7:M7")
            add_cell_value(
                worksheet4, 7, 10, 'Electons shared by atoms A and B', **tranparent_heading_properties)
            worksheet4.merge_cells("O7:R7")
            add_cell_value(
                worksheet4, 7, 15, 'Electrons delocalized by atom A to atom B', **tranparent_heading_properties)
            worksheet4.merge_cells("J6:R6")
            add_cell_value(
                worksheet4, 6, 10, 'Shared and delocalized electron counts are printed in the descending order', **colored_heading_properties)
            worksheet4.merge_cells("U6:AE6")
            add_cell_value(
                worksheet4, 6, 21, 'The number of electrons delocalized by atom A to atom B; these atoms are involved in the non-covalent intramolecular interactions', **tranparent_heading_properties)
            worksheet4.merge_cells(
                start_row=6 + cell_row2, start_column=21, end_row=6 + cell_row2, end_column=31)
            worksheet4.merge_cells(
                start_row=7 + cell_row2, start_column=21, end_row=7 + cell_row2, end_column=31)

            add_cell_value(worksheet4,  6+cell_row2, 21,
                           "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)
            add_cell_value(worksheet4,  7+cell_row2, 21,
                           "The number of electrons delocalized by atom A to atom B; these atoms are involved in the non-covalent intramolecular interactions", **tranparent_heading_properties)

            worksheet4.merge_cells("F2:N2")
            add_cell_value(
                worksheet4, 2, 6, "The minimum threshold number of shared electrons by atoms classically considered as being covalently bonded is:", **tranparent_heading_properties)
            add_cell_value(worksheet4, 2, 15, cov_threshold,
                           **heading_properties)

            worksheet4.merge_cells(start_row=cell_row3, start_column=21,
                                   end_row=cell_row3, end_column=31)
            add_cell_value(worksheet4, cell_row3, 21,
                           "The number of electrons delocalized to atom A by atom B; these atoms are involved in the non-covalent intramolecular interactions", **tranparent_heading_properties)

            worksheet4.merge_cells(start_row=cell_row4, start_column=21,
                                   end_row=cell_row4, end_column=31)
            add_cell_value(worksheet4, cell_row4, 21,
                           "The number of electrons delocalized to atom A by atom B; these atoms are involved in the non-covalent intramolecular interactions", **tranparent_heading_properties)
            worksheet4.merge_cells(start_row=cell_row4-1, start_column=21,
                                   end_row=cell_row4-1, end_column=31)
            add_cell_value(worksheet4, cell_row4-1, 21,
                           "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)

            for range_string in ["C3:R4", "J7:M7", "O7:R7", "U6:AE6", "U"+str(7+cell_row2)+":AE"+str(7+cell_row2), "U"+str(cell_row3)+":AE"+str(cell_row3), "U"+str(cell_row4)+":AE"+str(cell_row4)]:
                cells_range = worksheet4[range_string]
                for row in cells_range:
                    for cell in row:
                        cell.style = heading_style

            worksheet4["C3"].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)

            #                **heading_properties)
            # add_cell_value(worksheet4, 7, 5, 'e-shared', **heading_properties)
            # add_cell_value(worksheet4, 7, 12, 'e-shared', **heading_properties)
            # add_cell_value(worksheet4, 7, 16, 'e-delocalized by:',
            #                **heading_properties)
            # add_cell_value(
            #     worksheet4, 6, 10, 'Values of shared and delocalized electrons are in the descending order', **heading_properties)
            # add_cell_value(
            #     worksheet4, 6, 19, 'The number of electrons delocalized by atom A to atom B', **heading_properties)
            # add_cell_value(worksheet4, number_of_atoms+12, 19,
            #                'in the descending order', **heading_properties)
            temp_row = temp+3

            for i, name in enumerate(molecule_names):
                cell = worksheet4.cell(row=temp_row+i, column=21)
                cell.value = "Molecule-"+str(i+1)+" : "
                cell.font = Font(bold=True)
                cell = cell.offset(row=0, column=2)
                cell.value = name

            cell = worksheet4.cell(
                row=temp_row+2+number_of_molecules, column=21)
            cell.value = "Molecular system:"
            cell.font = Font(bold=True)
            cell = cell.offset(row=0, column=3)
            cell.value = system_name

            # add_cell_value(worksheet4, 8, 12, "e-shared", **heading_properties)
            # worksheet4.merge_cells("P8:Q8")
            # add_cell_value(worksheet4, 8, 16, "e-delocalized by:",
            #                **heading_properties)

            # worksheet4.merge_cells("B1:G1")

            # add_cell_value(
            #     worksheet4, 1, 2, "The threshold number of shared electrons by (A,B) is set to:", **mole_name_properties)
            # add_cell_value(worksheet4, 1, 8, cov_threshold,
            #                **heading_properties)

            # worksheet4.merge_cells("B3:K4")
            # add_cell_value(worksheet4, 3, 2, "The number of electrons (i) shared by atom-piar (A,B) (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A computed for all unique intramolecular atom-pairs (A,B) of a molecular system", **heading_properties)

            # worksheet4.merge_cells("B7:D7")
            # add_cell_value(worksheet4, 7, 2,
            #                "(A,B) non-covalently bonded", **heading_properties)

            # worksheet4.merge_cells("J7:P7")
            # add_cell_value(
            #     worksheet4, 7, 10, "Values of shared and delocalized electrons are in the descending order", **heading_properties)

            # worksheet4.merge_cells("U7:AB7")
            # add_cell_value(
            #     worksheet4, 7, 21, "The number of electrons delocalized by atom A to non-covalently-bonded with A atom B", **heading_properties)

            # worksheet4.merge_cells(
            #     start_row=6 + row_add, start_column=21, end_row=6 + row_add, end_column=29)
            # add_cell_value(worksheet4, 6 + row_add, 21,
            #                "The number of electrons delocalized by atom A to non-covalently-bonded with A atom B", **heading_properties)

            # worksheet4.merge_cells(
            #     start_row=7+row_add, start_column=21, end_row=7+row_add, end_column=23)
            # add_cell_value(worksheet4, 7+row_add, 21,
            #                "in the descending order", **heading_properties)

            if image_filename == "" or image_filename == None:
                pass
            else:
                insert_image(image_filename, worksheet4,
                             'U'+str(temp_row+4+number_of_molecules))

                # change the width of all the columns
            for col in worksheet4.columns:
                key = col[0].column_letter
                worksheet4.column_dimensions[key].width = 11.78

            # Create a workbook and add a worksheet.

            # print(molecule_names)


            reversed_sorted_data_of_table2_inter = {}

            if number_of_molecules > 1:
                doing_msg.set_progress(
                    0.90, "Working on inter-Mol(A,B) sheet....", speed=40)
                worksheet5 = workbook.create_sheet("inter-Mol(A,B)")

                # Create header cell  for intermolecular atom-pairs in worksheet5 table

                headers = ["(A,B) count", "Atom A",
                           "Atom B", "(A,B)", "A to B", "B to A"]
                headers_properties = {
                    'font': header_font,
                    'alignment': openpyxl.styles.Alignment(horizontal='center', vertical='center'),
                    "fill": header_fill
                }
                headers_properties_LEFT = {
                    'font': header_font,

                    "fill": header_fill
                }
                data_properties = {
                    "number_format": "0.0000",
                }

                row_idx = 7

                for idx, datas_dict in enumerate(inter_mol_dict_pair.items(), start=1):
                    mol1, mol2 = datas_dict[0].split("_")
                    # print(mol1, mol2)
                    datas = datas_dict[1]

                    add_cell_value(worksheet5, row_idx-1, 2,
                                   mol1+" with "+mol2, **mole_name_properties)
                    worksheet5.merge_cells(
                        start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
                    add_cell_value(
                        worksheet5, row_idx, 2, "Non-bonded atom-pairs (A,B)", **mole_name_properties)
                    worksheet5.merge_cells(
                        start_row=row_idx, start_column=6, end_row=row_idx, end_column=7)
                    add_cell_value(worksheet5, row_idx, 5,
                                   "e-shared", **heading_properties)
                    add_cell_value(worksheet5, row_idx, 6,
                                   "e-delocalized by:", **heading_properties)

                    for range_string in ["B"+str(row_idx)+":D"+str(row_idx), "F"+str(row_idx)+":G"+str(row_idx), "E"+str(row_idx)+":E"+str(row_idx)]:
                        cells_range = worksheet5[range_string]
                        for row in cells_range:
                            for cell in row:
                                cell.style = heading_style

                    row_idx += 1
                    for col_idx, header in enumerate(headers, start=2):
                        if col_idx < 5:
                            add_cell_value(worksheet5, row_idx, col_idx,
                                           header, **headers_properties_LEFT)
                        else:
                            add_cell_value(worksheet5, row_idx, col_idx,
                                           header, **headers_properties)
                    row_idx += 1
                    for _, value in enumerate(datas.items(), start=1):

                        A, B = self.get_atom_pair(value[0])

                        add_cell_value(worksheet5, row_idx, 2, _)
                        add_cell_value(worksheet5, row_idx, 3, A)
                        add_cell_value(worksheet5, row_idx, 4, B)

                        for col_idx, v in enumerate(value[1], start=5):
                            add_cell_value(worksheet5, row_idx, col_idx,
                                           v, **data_properties)
                        row_idx += 1

                    add_cell_value(worksheet5, row_idx, 4,
                                   "Total:", **total_properties)
                    add_cell_value(worksheet5, row_idx, 5,
                                   sum([i[1][0] for i in datas.items()]), **number_properties)
                    add_cell_value(worksheet5, row_idx, 6,
                                   sum([i[1][1] for i in datas.items()]), **number_properties)
                    add_cell_value(worksheet5, row_idx, 7,
                                   sum([i[1][2] for i in datas.items()]), **number_properties)
                    row_idx += 4

                # write the header row
                headers = ["(A,B) count", "Atom A", "Atom B",
                           "(A,B)", "", "(A,B) count", "Atom A", "Atom B", "A to B"]
                for col_idx, header in enumerate(headers):
                    if col_idx != 4:
                        if col_idx == 3 or col_idx == 8:
                            add_cell_value(worksheet5, 8, col_idx+9,
                                           header, **headers_properties)
                        else:
                            add_cell_value(worksheet5, 8, col_idx+9,
                                           header, **headers_properties_LEFT)
                        # add_cell_value(worksheet5, 9, col_idx+9,
                        #                header, **headers_properties)

                for row_idx, value in enumerate(inter_sorted_according_to_total_deloc_electron.items(), start=9):
                    A, B = self.get_atom_pair(value[0])

                    new_values = [int(row_idx-8), A, B, value[1]]

                    for col_idx, v in enumerate(new_values, start=9):
                        cell = worksheet5.cell(row=row_idx, column=col_idx)
                        cell.value = v
                        if col_idx != 9:
                            cell.number_format = "0.0000"

                for row_idx, value in enumerate(inter_sorted_according_to_contri.items(), start=9):
                    A, B = self.get_atom_pair(value[0])

                    new_values = [int(row_idx-8), A, B, value[1]]

                    for col_idx, v in enumerate(new_values, start=14):
                        cell = worksheet5.cell(row=row_idx, column=col_idx)
                        cell.value = v
                        if col_idx != 14:
                            cell.number_format = "0.0000"

                # write the footer row
                add_cell_value(worksheet5, len(
                    inter_sorted_according_to_total_deloc_electron)+9, 11, "Total:", **total_properties)
                add_cell_value(worksheet5, len(inter_sorted_according_to_total_deloc_electron)+9, 12, sum(
                    [i[1] for i in inter_sorted_according_to_total_deloc_electron.items()]), **number_properties)

                add_cell_value(worksheet5, len(
                    inter_sorted_according_to_contri)+9, 16, "Total:", **total_properties)
                add_cell_value(worksheet5, len(inter_sorted_according_to_contri)+9, 17, sum(
                    [i[1] for i in inter_sorted_according_to_contri.items()]), **number_properties)

                col_idx = 20
                temp = 0
                for key, value in data_of_table2_inter.items():

                    add_cell_value(worksheet5, 8, col_idx,
                                   "To atom B", **headers_properties_LEFT)
                    add_cell_value(worksheet5, 7, col_idx+1,
                                   "By atom A", **headers_properties)
                    add_cell_value(worksheet5, 8, col_idx+1,
                                   key, **headers_properties)

                    for row_idx, v in enumerate(value.items(), start=9):
                        add_cell_value(worksheet5, row_idx, col_idx,
                                       v[0], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       1, v[1], **data_properties)

                    add_cell_value(worksheet5, len(value.items())+9,
                                   col_idx, "Total:", **total_properties)
                    temp = max(temp, len(value.items())+9)
                    add_cell_value(worksheet5, len(value.items())+9, col_idx+1,
                                   sum([i[1] for i in value.items()]), **number_properties)

                    col_idx += 4

                col_idx = 20
                row_add = temp-3
                cell_row2 = row_add

                for key, value in sorted_data_of_table2_inter.items():

                    add_cell_value(worksheet5, row_add + 9, col_idx,
                                   "To atom B", **headers_properties_LEFT)
                    add_cell_value(worksheet5, row_add + 8, col_idx +
                                   1, "By atom A", **headers_properties)
                    add_cell_value(worksheet5, row_add + 9, col_idx +
                                   1, key, **headers_properties)
                    add_cell_value(worksheet5, row_add + 9, col_idx + 2,
                                   '''%-fraction''', **headers_properties)

                    tot = sum([i[1] for i in value.items()])
                    for row_idx, v in enumerate(value.items(), start=row_add + 10):

                        add_cell_value(worksheet5, row_idx, col_idx,
                                       v[0], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       1, v[1], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       2, v[1]*100/tot, **data_properties)

                    add_cell_value(worksheet5, row_add + len(value.items()) +
                                   10, col_idx, "Total:", **total_properties)
                    temp = max(temp, row_add + len(value.items())+10)
                    add_cell_value(worksheet5, row_add + len(value.items())+10, col_idx+1,
                                   sum([i[1] for i in value.items()]), **number_properties)
                    add_cell_value(
                        worksheet5, row_add + len(value.items())+10, col_idx+2, 100, **total_properties)
                    col_idx += 4

                for key, value in data_of_table2_inter.items():
                    for k, v in value.items():
                        if k not in reversed_data_of_table2_inter:
                            reversed_data_of_table2_inter[k] = {}
                        reversed_data_of_table2_inter[k][key] = v

                for key, value in reversed_data_of_table2_inter.items():
                    reversed_sorted_data_of_table2_inter[key] = dict(
                        sorted(value.items(), key=lambda item: item[1], reverse=True))

                col_idx = 20
                row_add = temp+25+number_of_molecules
                cell_row3 = row_add + 7
                for key, value in reversed_data_of_table2_inter.items():

                    add_cell_value(worksheet5, row_add + 9, col_idx,
                                   "By atom B", **headers_properties_LEFT)
                    add_cell_value(worksheet5, row_add + 8, col_idx +
                                   1, "To atom A", **headers_properties)
                    add_cell_value(worksheet5, row_add + 9, col_idx +
                                   1, key, **headers_properties)

                    tot = sum([i[1] for i in value.items()])
                    for row_idx, v in enumerate(value.items(), start=row_add + 10):

                        add_cell_value(worksheet5, row_idx, col_idx,
                                       v[0], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       1, v[1], **data_properties)

                    add_cell_value(worksheet5, row_add + len(value.items()) +
                                   10, col_idx, "Total:", **total_properties)
                    temp_ = max(temp, row_add + len(value.items())+10)
                    add_cell_value(worksheet5, row_add + len(value.items())+10, col_idx+1,
                                   sum([i[1] for i in value.items()]), **number_properties)
                    col_idx += 4

                col_idx = 20
                row_add = temp_-3
                cell_row4 = row_add + 7
                for key, value in reversed_sorted_data_of_table2_inter.items():

                    add_cell_value(worksheet5, row_add + 9, col_idx,
                                   "By atom B", **headers_properties_LEFT)
                    add_cell_value(worksheet5, row_add + 8, col_idx +
                                   1, "To atom A", **headers_properties)
                    add_cell_value(worksheet5, row_add + 9, col_idx +
                                   1, key, **headers_properties)
                    add_cell_value(worksheet5, row_add + 9, col_idx +
                                   2, '''%-fraction''', **headers_properties)

                    tot = sum([i[1] for i in value.items()])
                    for row_idx, v in enumerate(value.items(), start=row_add + 10):

                        add_cell_value(worksheet5, row_idx, col_idx,
                                       v[0], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       1, v[1], **data_properties)
                        add_cell_value(worksheet5, row_idx, col_idx +
                                       2, v[1]*100/tot, **data_properties)

                    add_cell_value(worksheet5, row_add + len(value.items()) +
                                   10, col_idx, "Total:", **total_properties)
                    temp_temp = max(temp, row_add + len(value.items())+10)
                    add_cell_value(worksheet5, row_add + len(value.items())+10, col_idx+1,
                                   sum([i[1] for i in value.items()]), **number_properties)
                    add_cell_value(
                        worksheet5, row_add + len(value.items())+10, col_idx+2, 100, **total_properties)
                    col_idx += 4

                worksheet5.insert_cols(1, 1)
                for merged_cell in worksheet5.merged_cells:
                    merged_cell.shift(1, 0)

                worksheet5.merge_cells("C3:R4")
                add_cell_value(worksheet5, 3, 3, '''The number of electrons (i) shared by atom-piar (A,B) is shown as 'e-shared', (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A are shown as 'e-delocalized by:'.                                                                                                      These values are computed for all unique intermolecular atom-pairs (A,B) of a molecular system. Typically, they are considered as being involved in non-covalent interactions.''', **tranparent_heading_properties)

                worksheet5.merge_cells("J7:M7")
                add_cell_value(
                    worksheet5, 7, 10, 'Electons shared by atoms A and B', **tranparent_heading_properties)
                worksheet5.merge_cells("O7:R7")
                add_cell_value(
                    worksheet5, 7, 15, 'Electrons delocalized by atom A to atom B', **tranparent_heading_properties)
                worksheet5.merge_cells("J6:R6")
                add_cell_value(
                    worksheet5, 6, 10, 'Shared and delocalized electron counts are printed in the descending order', **colored_heading_properties)
                worksheet5.merge_cells("U6:AE6")
                add_cell_value(
                    worksheet5, 6, 21, 'The number of electrons delocalized by atom A to atom B; these atoms are involved in the non-covalent intermolecular interactions', **tranparent_heading_properties)
                worksheet5.merge_cells(start_row=6+cell_row2,
                                       start_column=21, end_row=6+cell_row2, end_column=31)
                worksheet5.merge_cells(start_row=7+cell_row2,
                                       start_column=21, end_row=7+cell_row2, end_column=31)

                add_cell_value(worksheet5,  6+cell_row2, 21,
                               "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)
                add_cell_value(worksheet5,  7+cell_row2, 21,
                               "The number of electrons delocalized by atom A to atom B; these atoms are involved in the non-covalent intermolecular interactions", **tranparent_heading_properties)

                worksheet5.merge_cells("F2:N2")
                add_cell_value(
                    worksheet5, 2, 6, "The minimum threshold number of shared electrons by atoms classically considered as being covalently bonded is:", **tranparent_heading_properties)
                add_cell_value(worksheet5, 2, 15, cov_threshold,
                               **heading_properties)

                worksheet5.merge_cells(start_row=cell_row3, start_column=21,
                                       end_row=cell_row3, end_column=31)
                add_cell_value(worksheet5, cell_row3, 21,
                               "The number of electrons delocalized to atom A by each atom of a molecular system, atom B", **tranparent_heading_properties)
                worksheet5.merge_cells(start_row=cell_row4, start_column=21,
                                       end_row=cell_row4, end_column=31)
                add_cell_value(worksheet5, cell_row4, 21,
                               "The number of electrons delocalized to atom A by each atom of a molecular system, atom B", **tranparent_heading_properties)
                worksheet5.merge_cells(start_row=cell_row4-1, start_column=21,
                                       end_row=cell_row4-1, end_column=31)
                add_cell_value(worksheet5, cell_row4-1, 21,
                               "Values of delocalized electrons are printed in the descending order", **colored_heading_properties)

                for range_string in ["C3:R4", "J7:M7", "O7:R7", "U6:AE6", "U"+str(7+cell_row2)+":AE"+str(7+cell_row2), "U"+str(cell_row3)+":AE"+str(cell_row3), "U"+str(cell_row4)+":AE"+str(cell_row4)]:
                    cells_range = worksheet5[range_string]
                    for row in cells_range:
                        for cell in row:
                            cell.style = heading_style
                worksheet5["C3"].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)

                #                **heading_properties)
                # add_cell_value(worksheet5, 7, 5, 'e-shared', **heading_properties)
                # add_cell_value(worksheet5, 7, 12, 'e-shared', **heading_properties)
                # add_cell_value(worksheet5, 7, 16, 'e-delocalized by:',
                #                **heading_properties)
                # add_cell_value(
                #     worksheet5, 6, 10, 'Values of shared and delocalized electrons are in the descending order', **heading_properties)
                # add_cell_value(
                #     worksheet5, 6, 19, 'The number of electrons delocalized by atom A to atom B', **heading_properties)
                # add_cell_value(worksheet5, number_of_atoms+12, 19,
                #                'in the descending order', **heading_properties)

                for i, name in enumerate(molecule_names):
                    cell = worksheet5.cell(
                        row=temp+3+i, column=21)
                    cell.value = "Molecule-"+str(i+1)+" : "
                    cell.font = Font(bold=True)
                    cell = cell.offset(row=0, column=2)
                    cell.value = name

                cell = worksheet5.cell(
                    row=temp+5+number_of_molecules, column=21)
                cell.value = "Molecular system:"
                cell.font = Font(bold=True)
                cell = cell.offset(row=0, column=3)
                cell.value = system_name

                # add_cell_value(worksheet5, 8, 12, "e-shared",
                #                **heading_properties)
                # worksheet5.merge_cells("P8:Q8")
                # add_cell_value(worksheet5, 8, 16, "e-delocalized by:",
                #                **heading_properties)

                # worksheet5.merge_cells("B1:G1")
                # add_cell_value(
                #     worksheet5, 1, 2, "The minimum threshold number of shared electrons by atoms classically considered as being covalently bonded is:", **mole_name_properties)
                # add_cell_value(worksheet5, 1, 8, cov_threshold,
                #                **heading_properties)

                # worksheet5.merge_cells("B3:K4")
                # add_cell_value(worksheet5, 3, 2, "The number of electrons (i) shared by atom-piar (A,B) (ii) delocalized by atom A to atom B and (iii) delocalized by atom B to atom A computed for all unique intermolecular atom-pairs (A,B) of a molecular system", **heading_properties)

                # worksheet5.merge_cells("B7:G7")
                # add_cell_value(
                #     worksheet5, 7, 2, "(A,B) involved in non-covalent intermolecular interaction", **heading_properties)

                # worksheet5.merge_cells("J7:P7")
                # add_cell_value(
                #     worksheet5, 7, 10, "Values of shared and delocalized electrons are in the descending order", **heading_properties)

                # worksheet5.merge_cells("U7:AB7")
                # add_cell_value(
                #     worksheet5, 7, 21, "The number of electrons delocalized by atom A to atom B of another molecule", **heading_properties)

                # worksheet5.merge_cells(
                #     start_row=6 + row_add, start_column=21, end_row=6 + row_add, end_column=28)
                # add_cell_value(worksheet5, 6 + row_add, 21,
                #                "The number of electrons delocalized by atom A to atom B of another molecule", **heading_properties)

                # worksheet5.merge_cells(
                #     start_row=7+row_add, start_column=21, end_row=7+row_add, end_column=23)
                # add_cell_value(worksheet5, 7+row_add, 21,
                #                "in the descending order", **heading_properties)

                if image_filename == "" or image_filename == None:
                    pass
                else:
                    insert_image(image_filename, worksheet5,
                                 'U'+str(temp+5+number_of_molecules+2))

                    # change the width of all the columns
                for col in worksheet5.columns:
                    key = col[0].column_letter
                    worksheet5.column_dimensions[key].width = 11.78

            doing_msg.set_progress(1, "Finishing up....", speed=40)
            doing_msg.destroy()

        except Exception as e:

            # print(e)
            doing_msg.destroy()
            msg = CTkMessagebox(
                title="Error", message="Something went wrong while generating excel file!!!\n\n"+str(e), icon="cancel", master=app)
            return

    #   ENDDDDDDDDDDDDDDD

        # except Exception as e:
        #         print(e)
        #         msg = CTkMessagebox(title="Error", message="Something went wrong while generating excel file!!!\n\n"+str(e), icon="cancel")
        #         return

        msg = CTkMessagebox(title="Success!!", message="Data is successfully generated.\n",
                            icon="check", master=app)
        response = msg.get()

        # # Clears any existing text in entry widget
        # self.number_of_frag_entry.delete(0, END)
        # # Clears any existing text in entry widget
        # self.frag_input.delete("0.0", "end")

        try:

            save_file = filedialog.asksaveasfilename(
                initialdir=os.getcwd(),
                initialfile=system_name+"_MOWeD-LAC.xlsx",
                title="Save output file",
                filetype=[("Excel file", "*.xlsx")],
            )
            print("Saving file in: ", save_file)

            if not save_file:
                return

            _ = {
                "number_of_molecules": number_of_molecules,
                "molecule_relations": molecule_relations,
                "molecule_name": molecule_name,
                "molecule_names": molecule_names,
                "molecules_numbers": molecules_numbers,
                "number_of_atoms": number_of_atoms,
                "atoms_names": atoms_names,
                "total_electron": total_electron,
                "loc_electron": loc_electron,
                "deloc_electron": deloc_electron,
                "intra_molecule": intra_molecule,
                "inter_molecule": inter_molecule,
                "atoms_pairs": atoms_pairs,
                "total_deloc_electron": total_deloc_electron,
                "A_contri": A_contri,
                "B_contri": B_contri,
                "data_of_table2": data_of_table2,
                "sorted_data_of_table2": sorted_data_of_table2,
                "reversed_data_of_table2": reversed_data_of_table2,
                "reversed_sorted_data_of_table2": reversed_sorted_data_of_table2,
                "atom_molecule_dict": atom_molecule_dict,
                "reversed_data_of_table2_intra": reversed_data_of_table2_intra,
                "reversed_sorted_data_of_table2_intra": reversed_sorted_data_of_table2_intra,
                "reversed_data_of_table2_inter": reversed_data_of_table2_inter,
                "reversed_sorted_data_of_table2_inter": reversed_sorted_data_of_table2_inter,

                # "workbook": workbook,
                # "image_obj": image_obj,
                # "pil_image": pil_image,
                "system_name": system_name,
                "cov_threshold": cov_threshold,

                "sorted_data": sorted_data,
                "total_for_each_molecule": total_for_each_molecule,
                "intra_mol_data": intra_mol_data,
                "intra_mol_dict": intra_mol_dict,
                "intra_sorted_according_to_total_deloc_electron": intra_sorted_according_to_total_deloc_electron,
                "intra_sorted_according_to_contri": intra_sorted_according_to_contri,
                "data_of_table2_intra": data_of_table2_intra,
                "sorted_data_of_table2_intra": sorted_data_of_table2_intra,


                "inter_mol_dict": inter_mol_dict,
                "inter_mol_data": inter_mol_data,
                "inter_mol_dict_pair": inter_mol_dict_pair,
                "inter_sorted_according_to_total_deloc_electron": inter_sorted_according_to_total_deloc_electron,
                "inter_sorted_according_to_contri": inter_sorted_according_to_contri,
                "data_of_table2_inter": data_of_table2_inter,
                "sorted_data_of_table2_inter": sorted_data_of_table2_inter,

            }
            # pil_image.show()
            filename, ext = os.path.splitext(save_file)
            # add extension as .pickle
            filename_pickle = filename + ".pickle"

            # Store dictionaries in a pickle file
            with open(filename_pickle, "wb") as f:
                dump(_, f)

            workbook.save(save_file)

            msg = CTkMessagebox(
                title="Info", message="The output file is saved in the following path:\n\n"+save_file, icon="check", master=app)

            os.startfile(save_file)

        except Exception as e:
            print(e)
            msg = CTkMessagebox(
                title="Error", message="Something went wrong while saving the file!!!\n\n"+str(e), icon="cancel", master=app)
            return

        # # Clears any existing text in entry widget
        # self.num_mols_entry.delete(0, END)
        # # Clears any existing text in entry widget
        # self.cov_threshold_entry.delete("0.0", "end")
        # # Clears any existing text in entry widget
        # self.image_filename_entry.delete(0, END)
        self.container.home_frame = ATOMIC_HomeFrame(self.container.container, self.container)
        self.container.page2_frame = ATOMIC_Page2Frame(self.container.container, self.container)

        self.container.show_frame(self.container.home_frame)



    def insert_image(self, image_path, worksheet, coordinates):
        try:
            # Load image and calculate new dimensions to match required height
            img = Image(image_path)
            width, height = img.width, img.height
            aspect_ratio = width / height
            new_height = 377.946666667  # 10 cm in pixels (at 96 dpi)
            new_width = new_height * aspect_ratio
            img.height = new_height
            img.width = new_width

            # Insert the image into the specified cell
            cell = worksheet[coordinates]
            worksheet.add_image(img, coordinates)
        except:
            print("Failed to insert image: {}".format(image_path))

    def get_names_list(self, numbers, atom_names):
        names = []
        for num in numbers:
            names.append(atom_names[num-1])
        return names

    def get_numbers(self, string):
        """Parses comma-separated numbers and ranges into a list of integers."""
        try:
            numbers = []
            blocks = string.split(',')
            for block in blocks:
                # print(block)
                if '-' in block:
                    # print(block + "has -")
                    if len(block.split('-')) != 2:
                        raise Exception
                    start, end = map(int, block.split('-'))
                    # print(start, end)
                    numbers.extend(range(start, end + 1))
                    # print(numbers)
                else:
                    numbers.append(int(block))
            if len(numbers) != len(set(numbers)):
                raise Exception

            return sorted(numbers)
        except:
            raise Exception


class ATOMIC_Page2Frame(customtkinter.CTkFrame):
    def __init__(self, master, container, **kwargs):

        super().__init__(master, **kwargs)

        self.container = container

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure(4, weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="MOWeD-LAC", font=customtkinter.CTkFont(family="Candara",size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.appearance_mode_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=[
                                                                       "Light", "Dark", "System"], command=self.container.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="GUI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%", "130%", "140%", "150%"],
                                                               command=self.container.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # Default value

        self.appearance_mode_optionemenu.set(
            customtkinter.get_appearance_mode())
        self.scaling_optionemenu.set("100%")

        self.num_processor_label = customtkinter.CTkLabel(
            self, text="Number of Processors:", anchor="w")
        self.num_processor_label.grid(row=0, column=1, padx=20, pady=(
            20, 20), sticky="nsew", columnspan=2)

        self.num_processor_entry = CustomSpinbox(
            self, width=150, step_size=1, data_type=int)
        self.num_processor_entry.grid(
            row=0, column=3, padx=20, pady=(20, 20), sticky="nsew")
        self.num_processor_entry.set(4)

        self.mem_limit_label = customtkinter.CTkLabel(
            self, text="Memory Limit (MB):", anchor="w")
        self.mem_limit_label.grid(row=1, column=1, padx=20, pady=(
            20, 20), sticky="nsew", columnspan=2)
        self.mem_limit_entry = CustomSpinbox(
            self, width=150, step_size=1000, data_type=int)
        self.mem_limit_entry.grid(
            row=1, column=3, padx=20, pady=(20, 20), sticky="nsew")
        self.mem_limit_entry.set(1000)

        # self.wfx_file_label = customtkinter.CTkLabel(
        #     self, text="Choose WFX File:", anchor="w")
        # self.wfx_file_label.grid(row=4, column=1,columnspan=2, padx=20, pady=(0, 0), sticky="w")

        self.wfx_file_input = customtkinter.CTkEntry(
            self, placeholder_text="Enter WFX File Path", width=150)
        self.wfx_file_input.grid(
            row=2, column=1, columnspan=2, padx=20, pady=(20, 0), sticky="nsew")

        self.select_wfx_file_button = customtkinter.CTkButton(
            self, text="Browse", command=self.select_wfx_file_event)
        self.select_wfx_file_button.grid(
            row=2, column=3, padx=20, pady=(20, 0), sticky="nsew")


        self.left_bottom_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.left_bottom_frame.grid(
            row=5, column=1, padx=20, pady=(0, 20), sticky="nsew")

        self.help_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2753", command=self.open_help, border_width=0, width=30,height=29)
        self.help_button.grid(row=0, column=0, padx=(0,10),
                                pady=0)



        self.back_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2190", command=lambda: self.back_to_home(), width=30, height=30)
        self.back_button.grid(row=0, column=1, padx=(10,0),
                              pady=0)

        self.submit_button = customtkinter.CTkButton(
            self, text="Submit", command=self.submit_event)
        self.submit_button.grid(row=5, column=3, padx=20,
                                pady=(0, 20), sticky="nsew")

    def open_help(self):
        self.container.open_help(1)

    def back_to_home(self):
        self.container.show_frame(self.container.home_frame)

    def select_wfx_file_event(self):
        # Ask the user to select the input file

        

        self.input_file = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            title="Select input file",
            filetypes=[("WFX files", "*.wfx")]
        )
        if not self.input_file:
            print("No file selected")
            return

        # Split the dir and file name
        dir_name, file_name_of_wxf = os.path.split(self.input_file)
        # Update the input field
        self.wfx_file_input.delete(0, END)
        os.chdir(dir_name)

        self.wfx_file_input.insert(0, file_name_of_wxf)
        print("Dir name: " + dir_name)
        print("File name: " + file_name_of_wxf)

    def submit_event(self):
        try:
            num_processor = int(self.num_processor_entry.get())
            mem_limit = int(self.mem_limit_entry.get())
            wfx_file = self.wfx_file_input.get()
        except Exception as e:
            print(e)
            msg = CTkMessagebox(
                title="Error", message="Something went wrong\n"+str(e), icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return

        if not num_processor:
            msg = CTkMessagebox(
                title="Error", message="Please enter the number of processors", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        if not mem_limit:
            msg = CTkMessagebox(
                title="Error", message="Please enter the memory limit", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        if not wfx_file:
            msg = CTkMessagebox(
                title="Error", message="Please select the wfx file", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return

        # print("Number of Processor: " + num_processor)
        # print("Memory Limit: " + mem_limit)
        # print("WFX File: " + wfx_file)


        fixed_command = "N " + str(num_processor) + "\n" + "WFX " + str(wfx_file) + "\n" + "MEM " + str(mem_limit) + "\n"
        addiional_command = "COORD CUBE" + "\n" + "OVERLAP LDO" + "\n"
        if self.container.faldi_commands and self.container.faldi_commands != "":
            # Remove all trailing and leading spaces
            addiional_command = self.container.faldi_commands
            print("Additional command: " + addiional_command[0])

        # create a inp.add file and write the input to it
        inp_add_file = open(os.path.join(os.getcwd(), "inp.add"), "w")
        inp_add_file.write(fixed_command + addiional_command)
        inp_add_file.close()

        print("inp.add file created")
        default_exe_file = os.path.join(base_path, 'assets', "FALDI_pops.exe")

        if not os.path.exists(default_exe_file):
            print("Executable file does not exist")
            return
        else:
            print("Executable file exists")

        print("Running the FALDI-pops.exe file with the inp.add file")

# import os

        # Set the path of the executable file and input file

        input_file = "inp.add"

        # Run the executable file with the provided arguments
        cmd = "\""+default_exe_file+"\"" + " " + input_file
        print(cmd)

        wait_msg = ProgressCTkMessagebox(
            title="Please Wait", message="FALDI is running", icon="info", fade_in_duration=1, master=app)
        wait_msg.set_progress(0)
        
        try:
            os.system(cmd)
        except Exception as e:
            print(e)
            msg = CTkMessagebox(
                title="Error", message="Something went while running FALDI!!\n"+str(e), icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        wait_msg.set_progress(1, "FALDI-pops.exe file finished running", speed=40)
        wait_msg.destroy()

        print("FALDI-pops.exe file finished running")

        print("Reading the FALDI-pops.sum file")
        try:
            # Read the data from the selected file
            with open("FALDI-pops.sum", 'r') as sum_file:
                sum_data = sum_file.readlines()

            # Split each line of the SUM file by spaces to separate the columns
            split_data = [line.split() for line in sum_data]
            print("FALDI-pops.sum file read successfully")

            # create a new workbook object
            workbook = openpyxl.Workbook()

            # select the active worksheet
            worksheet = workbook.active

            # write the data to the worksheet
            for row in split_data:
                worksheet.append(row)

            # Ask the user to select where to save the output file
            # save_filename = "example.xlsx"

            # save the workbook to a file
            # workbook.save(save_filename)

            # Find the index of the row that separates the two tables
            for i, row in enumerate(worksheet.iter_rows()):
                if all(cell.value is None for cell in row):
                    separator_index = i
                    break

            # Split the table vertically into two separate tables
            table1 = worksheet.iter_rows(
                min_row=2, max_row=separator_index, values_only=True)
            table2 = worksheet.iter_rows(
                min_row=separator_index+3, values_only=True)

            # Extract data from table 1
            atom_col_idx = 0
            n_col_idx = 1
            li_col_idx = 2
            tot_di_col_idx = 3

            atoms_names = []
            total_electron = []
            loc_electron = []
            deloc_electron = []

            for row in table1:
                atoms_names.append(row[atom_col_idx])
                total_electron.append(float(row[n_col_idx]))
                loc_electron.append(float(row[li_col_idx]))
                deloc_electron.append(float(row[tot_di_col_idx]))

            # print(atoms_names, total_electron, loc_electron, deloc_electron)

            number_of_atoms = len(atoms_names)

            # create empty lists of equal length to atoms_names
            molecule_name = [''] * (len(atoms_names))
            intra_molecule = [0.0] * len(atoms_names)
            inter_molecule = [0.0] * len(atoms_names)

            # Extract data from table 2
            atoms_pairs = []
            total_deloc_electron = []
            A_contri = []
            B_contri = []

            for row in table2:
                atoms_pairs.append(row[atom_col_idx])
                total_deloc_electron.append(float(row[n_col_idx]))
                A_contri.append(float(row[li_col_idx]))
                B_contri.append(float(row[tot_di_col_idx]))

            # print(atoms_pairs, total_deloc_electron, A_contri, B_contri)

            # # read in the data from the file
            # df = pd.read_excel('example.xlsx')

            # # find the index of the row that separates the two tables
            # separator_index = df.index[df.isnull().all(axis=1)][0]

            # # split the table vertically into two separate tables
            # table1 = df.iloc[:separator_index, :]
            # table2 = df.iloc[separator_index+2:, :]

            # # extract data from table 1
            # atoms_names = table1['Atom'].tolist()
            # total_electron = table1['N'].astype(float).tolist()
            # loc_electron = table1['LI'].astype(float).tolist()
            # deloc_electron = table1['Tot_DI'].astype(float).tolist()

            # number_of_atoms = len(atoms_names)

            # # create empty lists of equal length to atoms_names
            # molecule_name = [''] * (len(atoms_names)+1)
            # intra_molecule = [0.0] * len(atoms_names)
            # inter_molecule = [0.0] * len(atoms_names)

            # # extract data from table 2
            # atoms_pairs = table2['Atom'].tolist()
            # total_deloc_electron = table2['N'].astype(float).tolist()
            # A_contri = table2['LI'].astype(float).tolist()
            # B_contri = table2['Tot_DI'].astype(float).tolist()

            print("Parsing SUM file...")
            print("SUM file has been parsed successfully...")

            _ = {
                "number_of_atoms": number_of_atoms,
                "atoms_names": atoms_names,
                "total_electron": total_electron,
                "loc_electron": loc_electron,
                "deloc_electron": deloc_electron,
                "molecule_name": molecule_name,
                "intra_molecule": intra_molecule,
                "inter_molecule": inter_molecule,
                "atoms_pairs": atoms_pairs,
                "total_deloc_electron": total_deloc_electron,
                "A_contri": A_contri,
                "B_contri": B_contri,
                "image_filename": ""
            }

            # print(_)

            self.container.data_for_this_file = _

        except Exception as e:
            msg = CTkMessagebox(
                title="Warning!!", message="Error parsing SUM file\n"+str(e),
                icon="warning", master=app)
            if msg.get() == "OK":
                pass
            return

        # self.controller.destroy()
        # self.container.show_frame(self.container.input_frame)
        text_to_display = ""

        text_to_display += "No of atoms: \t" + str(number_of_atoms) + "\n"
        text_to_display += "Atom names: \t"
        for idx, atom_name in enumerate(atoms_names):
            text_to_display += atom_name + ', '
        #     if idx % 8 == 0 and idx != 0:
        #         text_to_display += "\n\t\t"
        # text_to_display += "\n\n"

        self.container.input_frame.info_label.configure(state="normal")
        self.container.input_frame.info_label.delete("1.0", "end")
        self.container.input_frame.info_label.insert("end", text_to_display)
        self.container.input_frame.info_label.configure(state="disabled")

        self.container.input_frame.from_window = "page2_frame"
        self.container.show_frame(self.container.input_frame)



        # msg = CTkMessagebox(
        #     title="Info about the extracted data!", message=text_to_display, icon=None,  width=500, button_width=120, button_height=30, widen=1)
        # if msg.get() == "OK":
        #     pass



class ScrollableCheckBoxFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, item_list, command=None, **kwargs):
        super().__init__(master, **kwargs)

        self.command = command
        self.checkbox_list = []
        for i, item in enumerate(item_list):
            self.add_item(item)

    def add_item(self, item):
        checkbox = customtkinter.CTkCheckBox(self, text=item)
        if self.command is not None:
            checkbox.configure(command=self.command)
        checkbox.grid(row=len(self.checkbox_list), column=0, pady=(0, 10))
        self.checkbox_list.append(checkbox)

    def remove_item(self, item):
        for checkbox in self.checkbox_list:
            if item == checkbox.cget("text"):
                checkbox.destroy()
                self.checkbox_list.remove(checkbox)
                return

    def get_checked_items(self):
        return [checkbox.cget("text") for checkbox in self.checkbox_list if checkbox.get() == 1]


class ATOMIC_Page3Frame(customtkinter.CTkFrame):
    def __init__(self, master, container, **kwargs):

        super().__init__(master, **kwargs)

        self.container = container

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure(4, weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(
            self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=6, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="MOWeD-LAC", font=customtkinter.CTkFont(family="Candara",size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))

        self.appearance_mode_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=[
                                                                       "Light", "Dark", "System"], command=self.container.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(
            row=6, column=0, padx=20, pady=(10, 10))

        self.scaling_label = customtkinter.CTkLabel(
            self.sidebar_frame, text="GUI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%", "130%", "140%", "150%"],
                                                               command=self.container.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # Default value

        self.appearance_mode_optionemenu.set(
            customtkinter.get_appearance_mode())
        self.scaling_optionemenu.set("100%")

        self.num_processor_label = customtkinter.CTkLabel(
            self, text="Number of Processors:", anchor="w")
        self.num_processor_label.grid(row=0, column=1, padx=20, pady=(
            20, 0), sticky="nsew", columnspan=2)

        self.num_processor_entry = CustomSpinbox(
            self, width=150, step_size=1, data_type=int)
        self.num_processor_entry.grid(
            row=0, column=3, padx=20, pady=(20, 0), sticky="nsew")
        self.num_processor_entry.set(4)

        self.mem_limit_label = customtkinter.CTkLabel(
            self, text="Memory Limit (MB):", anchor="w")
        self.mem_limit_label.grid(row=1, column=1, padx=20, pady=(
            20, 0), sticky="nsew", columnspan=2)
        self.mem_limit_entry = CustomSpinbox(
            self, width=150, step_size=1000, data_type=int)
        self.mem_limit_entry.grid(
            row=1, column=3, padx=20, pady=(20, 0), sticky="nsew")
        self.mem_limit_entry.set(1000)

        # self.wfx_file_label = customtkinter.CTkLabel(
        #     self, text="Choose WFX File:", anchor="w")
        # self.wfx_file_label.grid(row=4, column=1,columnspan=2, padx=20, pady=(0, 0), sticky="w")

        self.wfx_file_input = customtkinter.CTkEntry(
            self, placeholder_text="Enter WFX File Path", width=150)
        self.wfx_file_input.grid(
            row=2, column=1, columnspan=2, padx=20, pady=(20, 0), sticky="nsew")

        self.select_wfx_file_button = customtkinter.CTkButton(
            self, text="Browse", command=self.select_wfx_file_event)
        self.select_wfx_file_button.grid(
            row=2, column=3, padx=20, pady=(20, 0), sticky="nsew")
        
        self.image_filename_entry = customtkinter.CTkEntry(
            self, placeholder_text="Enter the image filename")
        self.image_filename_entry.grid(
            row=3, column=1, columnspan=2, padx=(20, 20), pady=(20, 10), sticky="nsew")

        self.browse_image_button = customtkinter.CTkButton(
            self, text="Browse",  command=lambda: self.handle_browse_image_button(self.image_filename_entry))
        self.browse_image_button.grid(row=3, column=3, padx=(
            20, 20), pady=(20, 10), sticky="nsew")
        
        
        
        

        
        
        
        


        self.left_bottom_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.left_bottom_frame.grid(
            row=5, column=1, padx=20, pady=(0, 20), sticky="nsew")

        self.help_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2753", command=self.open_help, border_width=0, width=30,height=29)
        self.help_button.grid(row=0, column=0, padx=(0,10),
                                pady=0)



        self.back_button = customtkinter.CTkButton(
            self.left_bottom_frame, text="\u2190", command=lambda: self.back_to_home(), width=30, height=30)
        self.back_button.grid(row=0, column=1, padx=(10,0),
                              pady=0)

        self.submit_button = customtkinter.CTkButton(
            self, text="Submit", command=self.submit_event)
        self.submit_button.grid(row=5, column=3, padx=20,
                                pady=(0, 20), sticky="nsew")
        
        
        
    def handle_browse_image_button(self, entry: customtkinter.CTkEntry):

        image_filename = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            title="Select an image",
            filetypes=[("PNG files", "*.png"), ("JPEG files", "*.jpg")]
        )

        if image_filename:
            try:

                file_dir, file_name = os.path.split(image_filename)
                print("Image filename: ", file_name)
                os.startfile(image_filename)
                entry.delete(0, "end")
                entry.insert(0, file_name)

            except:
                msg = CTkMessagebox(
                    title="Error", message="An error occured while loading the image!", icon="cancel", master=app)
                if msg.get() == "OK":
                    pass
                return

        else:
           
            msg = CTkMessagebox(
                title="Error", message="Please select an image!", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
        

    def open_help(self):
        self.container.open_help(1)

    def back_to_home(self):
        self.container.show_frame(self.container.home_frame)

    def select_wfx_file_event(self):
        # Ask the user to select the input file

        

        self.input_file = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            title="Select input file",
            filetypes=[("WFX files", "*.wfx")]
        )
        if not self.input_file:
            print("No file selected")
            return

        # Split the dir and file name
        dir_name, file_name_of_wxf = os.path.split(self.input_file)
        # Update the input field
        self.wfx_file_input.delete(0, END)
        os.chdir(dir_name)

        self.wfx_file_input.insert(0, file_name_of_wxf)
        print("Dir name: " + dir_name)
        print("File name: " + file_name_of_wxf)
        
        
        # Search for .sum file in the same directory
        for file in os.listdir(dir_name):
            if file.endswith(".sum"):
                self.sum_file = os.path.join(dir_name, file)
                print("SUM file found: " + self.sum_file)
                break
            
        
        
        with open(self.sum_file, 'r') as sum_file:
            sum_data = sum_file.readlines()
        # Split each line of the SUM file by spaces to separate the columns
        split_data = [line.split() for line in sum_data]

        print("FALDI-pops.sum file read successfully")

        # create a new workbook object
        workbook = openpyxl.Workbook()

        # select the active worksheet
        worksheet = workbook.active

        # write the data to the worksheet
        for row in split_data:
            worksheet.append(row)

        # # Ask the user to select where to save the output file
        save_filename = "example.xlsx"
        
        # Find the index of the row that separates the two tables
        for i, row in enumerate(worksheet.iter_rows()):
            if all(cell.value is None for cell in row):
                separator_index = i
                break


        table2 = worksheet.iter_rows(
            min_row=separator_index+3, values_only=True)

        # Extract data from table 1
        atom_col_idx = 0
        # print(atoms_names, total_electron, loc_electron, deloc_electron)

        self.atoms_pairs = []
       

        for row in table2:
            self.atoms_pairs.append(row[atom_col_idx])
            
        self.scrollable_checkbox_frame = ScrollableCheckBoxFrame(master=self, width=200, command=self.checkbox_frame_event,
                                                                 item_list=[atom_pair for atom_pair in self.atoms_pairs])
        
        self.scrollable_checkbox_frame.grid(row=4, column=1, columnspan=3, padx=20, pady=(20, 20), sticky="nsew")
        
        
        
        
        
        
        
        # self.scrollable_checkbox_frame.grid_rowconfigure(4, weight=1)
        
    def checkbox_frame_event(self):
        print(self.scrollable_checkbox_frame.get_checked_items())
        

    def submit_event(self):
        try:
            num_processor = int(self.num_processor_entry.get())
            mem_limit = int(self.mem_limit_entry.get())
            wfx_file = self.wfx_file_input.get()
            
            
        except Exception as e:
            print(e)
            msg = CTkMessagebox(
                title="Error", message="Something went wrong\n"+str(e), icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return

        if not num_processor:
            msg = CTkMessagebox(
                title="Error", message="Please enter the number of processors", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        if not mem_limit:
            msg = CTkMessagebox(
                title="Error", message="Please enter the memory limit", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        if not wfx_file:
            msg = CTkMessagebox(
                title="Error", message="Please select the wfx file", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        
        selected = self.scrollable_checkbox_frame.get_checked_items()
        if not selected:
            msg = CTkMessagebox(
                title="Error", message="Please select the atoms pairs", icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return

        print("Number of Processor: " + str(num_processor))
        print("Memory Limit: " + str(mem_limit))
        print("WFX File: " + wfx_file)

        unique_atoms_in_pairs = []
        for pair in selected:
            atoms = pair.split('_')
            unique_atoms_in_pairs.extend(atoms)
            
        unique_atoms_in_pairs = sorted(list(set(unique_atoms_in_pairs)))
        print("Unique atoms in pairs: ", unique_atoms_in_pairs)
        # return 

        

        fixed_command = "N " + str(num_processor) + "\n" + "WFX " + str(wfx_file) + "\n" + "MEM " + str(mem_limit) + "\n"
        addiional_command = "COORD CUBE" + "\n" + "OVERLAP LDO" + "\n" + "PRINt ATOMS"+ "\n"
        
        atoms_command = "SELECTED {0}\n{1}\n".format(len(unique_atoms_in_pairs), '\n'.join(unique_atoms_in_pairs))
        
        print(fixed_command + addiional_command + atoms_command)
        


        # return
        
        

        # create a inp.add file and write the input to it
        inp_add_file = open(os.path.join(os.getcwd(), "inp.add"), "w")
        inp_add_file.write(fixed_command + addiional_command + atoms_command)
        inp_add_file.close()

        print("inp.add file created")
        default_exe_file = os.path.join(base_path, 'assets', "FALDI_grid_12_02_2024.exe")

        if not os.path.exists(default_exe_file):
            print("Executable file does not exist")
            return
        else:
            print("Executable file exists")

        print("Running the FALDI-pops.exe file with the inp.add file")

# import os

        # Set the path of the executable file and input file

        input_file = "inp.add"

        # Run the executable file with the provided arguments
        cmd = "\""+default_exe_file+"\"" + " " + input_file
        print(cmd)

        wait_msg = ProgressCTkMessagebox(
            title="Please Wait", message="FALDI grid is running", icon="info", fade_in_duration=1, master=app)
        wait_msg.set_progress(0)
        
        try:
            os.system(cmd)
        except Exception as e:
            print(e)
            msg = CTkMessagebox(
                title="Error", message="Something went while running FALDI!!\n"+str(e), icon="cancel", master=app)
            if msg.get() == "OK":
                pass
            return
        wait_msg.set_progress(1, "FALDI-pops.exe file finished running", speed=40)
        wait_msg.destroy()

        print("FALDI-grid.exe file finished running")

        return

        print("Reading the FALDI-grid.sum file")
        try:
            # Read the data from the selected file
            with open("FALDI-pops.sum", 'r') as sum_file:
                sum_data = sum_file.readlines()

            # Split each line of the SUM file by spaces to separate the columns
            split_data = [line.split() for line in sum_data]
            print("FALDI-pops.sum file read successfully")

            # create a new workbook object
            workbook = openpyxl.Workbook()

            # select the active worksheet
            worksheet = workbook.active

            # write the data to the worksheet
            for row in split_data:
                worksheet.append(row)

            # Ask the user to select where to save the output file
            # save_filename = "example.xlsx"

            # save the workbook to a file
            # workbook.save(save_filename)

            # Find the index of the row that separates the two tables
            for i, row in enumerate(worksheet.iter_rows()):
                if all(cell.value is None for cell in row):
                    separator_index = i
                    break

            # Split the table vertically into two separate tables
            table1 = worksheet.iter_rows(
                min_row=2, max_row=separator_index, values_only=True)
            table2 = worksheet.iter_rows(
                min_row=separator_index+3, values_only=True)

            # Extract data from table 1
            atom_col_idx = 0
            n_col_idx = 1
            li_col_idx = 2
            tot_di_col_idx = 3

            atoms_names = []
            total_electron = []
            loc_electron = []
            deloc_electron = []

            for row in table1:
                atoms_names.append(row[atom_col_idx])
                total_electron.append(float(row[n_col_idx]))
                loc_electron.append(float(row[li_col_idx]))
                deloc_electron.append(float(row[tot_di_col_idx]))

            # print(atoms_names, total_electron, loc_electron, deloc_electron)

            number_of_atoms = len(atoms_names)

            # create empty lists of equal length to atoms_names
            molecule_name = [''] * (len(atoms_names))
            intra_molecule = [0.0] * len(atoms_names)
            inter_molecule = [0.0] * len(atoms_names)

            # Extract data from table 2
            atoms_pairs = []
            total_deloc_electron = []
            A_contri = []
            B_contri = []

            for row in table2:
                atoms_pairs.append(row[atom_col_idx])
                total_deloc_electron.append(float(row[n_col_idx]))
                A_contri.append(float(row[li_col_idx]))
                B_contri.append(float(row[tot_di_col_idx]))

            # print(atoms_pairs, total_deloc_electron, A_contri, B_contri)

            # # read in the data from the file
            # df = pd.read_excel('example.xlsx')

            # # find the index of the row that separates the two tables
            # separator_index = df.index[df.isnull().all(axis=1)][0]

            # # split the table vertically into two separate tables
            # table1 = df.iloc[:separator_index, :]
            # table2 = df.iloc[separator_index+2:, :]

            # # extract data from table 1
            # atoms_names = table1['Atom'].tolist()
            # total_electron = table1['N'].astype(float).tolist()
            # loc_electron = table1['LI'].astype(float).tolist()
            # deloc_electron = table1['Tot_DI'].astype(float).tolist()

            # number_of_atoms = len(atoms_names)

            # # create empty lists of equal length to atoms_names
            # molecule_name = [''] * (len(atoms_names)+1)
            # intra_molecule = [0.0] * len(atoms_names)
            # inter_molecule = [0.0] * len(atoms_names)

            # # extract data from table 2
            # atoms_pairs = table2['Atom'].tolist()
            # total_deloc_electron = table2['N'].astype(float).tolist()
            # A_contri = table2['LI'].astype(float).tolist()
            # B_contri = table2['Tot_DI'].astype(float).tolist()

            print("Parsing SUM file...")
            print("SUM file has been parsed successfully...")

            _ = {
                "number_of_atoms": number_of_atoms,
                "atoms_names": atoms_names,
                "total_electron": total_electron,
                "loc_electron": loc_electron,
                "deloc_electron": deloc_electron,
                "molecule_name": molecule_name,
                "intra_molecule": intra_molecule,
                "inter_molecule": inter_molecule,
                "atoms_pairs": atoms_pairs,
                "total_deloc_electron": total_deloc_electron,
                "A_contri": A_contri,
                "B_contri": B_contri,
                "image_filename": ""
            }

            # print(_)

            self.container.data_for_this_file = _

        except Exception as e:
            msg = CTkMessagebox(
                title="Warning!!", message="Error parsing SUM file\n"+str(e),
                icon="warning", master=app)
            if msg.get() == "OK":
                pass
            return

        # self.controller.destroy()
        # self.container.show_frame(self.container.input_frame)
        text_to_display = ""

        text_to_display += "No of atoms: \t" + str(number_of_atoms) + "\n"
        text_to_display += "Atom names: \t"
        for idx, atom_name in enumerate(atoms_names):
            text_to_display += atom_name + ', '
        #     if idx % 8 == 0 and idx != 0:
        #         text_to_display += "\n\t\t"
        # text_to_display += "\n\n"

        self.container.input_frame.info_label.configure(state="normal")
        self.container.input_frame.info_label.delete("1.0", "end")
        self.container.input_frame.info_label.insert("end", text_to_display)
        self.container.input_frame.info_label.configure(state="disabled")

        self.container.input_frame.from_window = "page2_frame"
        self.container.show_frame(self.container.input_frame)



        # msg = CTkMessagebox(
        #     title="Info about the extracted data!", message=text_to_display, icon=None,  width=500, button_width=120, button_height=30, widen=1)
        # if msg.get() == "OK":
        #     pass





# A-NH2	10-12
# A-F1	1-9,13-15
# A-CN	1,6
# A-F3	2-5,7-15
# A-CH	2,13
# A-F5	1,3-12,14,15
# A-rem	3-5,7-9,14,15
# A-F7	1,2,6,10-13
# A	1-15
# T	16-30
if __name__ == "__main__":
    app = ATOMIC_App()
    if getattr(sys, 'frozen', False):
        pyi_splash.close()
    app.mainloop()


'''
variabels:





number_of_molecules
molecule_relations
molecule_names
molecules_numbers
atoms_names
total_electron
loc_electron
deloc_electron
intra_molecule
inter_molecule
number_of_atoms
atoms_pairs
total_deloc_electron
A_contri
B_contri
data_of_table2
sorted_data_of_table2

N1H3	1,11,12,29
C5H2	5,13,14
C6H2	6,15,16
N2H	    2,17
C7H2	7,18,19
C8H2	8,20,21
N3H	    3,22
C9H2	9,23,24
C10H2	10,25,26
N4H2	4,27,28
'''
